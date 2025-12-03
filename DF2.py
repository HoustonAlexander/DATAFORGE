import tkinter as tk
import xlrd
import os
import xlwings as xw
import re
import fitz
import webbrowser
import pyautogui
import subprocess
import sys
import openpyxl
from pathlib import Path
import win32com.client
from win32com.client import makepy, Dispatch
from win32com.client import constants as c
import pythoncom
import numpy as np
from dateutil.relativedelta import relativedelta
from tkinter import ttk, BOTH, Menu, Text, messagebox
from datetime import date , datetime
from tkinter import filedialog
import sv_ttk
import threading
import csv
import calendar
import pandas as pd
import math, bisect
from tkinter import Label
from tkinter import Listbox
from tkinter import END
from tkinter import *
from PIL import ImageTk, Image
from openpyxl import load_workbook, Workbook
from urllib.parse import urlparse
import time
import itertools
from itertools import zip_longest
def donothing():
   x = 0

root = tk.Tk()
root.iconbitmap("assets/Anvil.ico")
root.tk.call('source','Azure/azure.tcl')
root.tk.call('set_theme', 'dark')
# file_path = tk.StringVar()
root.title("DATAFORGE2")
root.focus()
frame = ttk.Frame(root)
frame.columnconfigure(0, weight=1)
frame.pack()

menubar = Menu(root)
root.config(menu=menubar)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="New", command=donothing)
filemenu.add_command(label="Load", command=donothing)
filemenu.add_command(label="Save", command=donothing)
filemenu.add_separator()
filemenu.add_command(label="Exit", command=root.quit)
menubar.add_cascade(label="File", menu=filemenu)

helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Help Index", command=donothing)
helpmenu.add_command(label="About...", command=donothing)
menubar.add_cascade(label="Help", menu=helpmenu)
pd.set_option('display.min_rows', 56)
file_path = tk.StringVar()
file_path = "No File Found"

file_path2 = tk.StringVar()
file_path2 = "No File Found"

Reports = "-Choose Your Process-","The Number", "WireTap", "Closeout","RetroActivity", "PDFReader","Invictus", "AZAPPER", "QuickFormat", "AutoDraw - USDA NIFA", "AutoDraw - USDA NIFA", "AutoDraw - NSF", "P4: Payroller", "ASAP Drawsheet Formatter"
PDFReports = "-Choose Your Process-", "WireTap","RetroActivity", "PDFReader", "AZAPPER" #, "Invictus",
XLXSReports = "-Choose Your Process-", "AutoDraw - USDA NIFA", "AutoDraw - NIH", "AutoDraw - NSF", "P4: Payroller",  # "ASAP Drawsheet Formatter", , "The Number"

Letsgo = tk.StringVar()
today_date = datetime.today().strftime("%m_%d_%y")
# default_name = tk.StringVar()
defaultlo = os.path.join(os.path.expanduser("~"), "Documents")

dflon = Path(defaultlo).name
default_name = f"CNGWF{today_date}"
filecheck2 = default_name + ".csv"

print(filecheck2)
global file_types
file_types = []

def start_snip_thread():
   # selected_option = Letsgo.get()
   # SelectedFile = Path(FileSelection)
   sthread = threading.Thread(target=launch_snipping_tool)
   sthread.start()
   
def GetHelp():
   webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE")

def resource_path(*parts) -> Path:
    # In onefile mode PyInstaller unpacks to a temp dir in sys._MEIPASS
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base.joinpath(*parts)

def Consolidate():
   
   try:
      root = os.path.dirname(os.path.abspath(__file__))
      macro_path = resource_path("assets", "DataForgeMACROtmp.xlsm")
      print(macro_path)
      ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
      ExcelApp.visible = True
      Blank = ExcelApp.Workbooks.Add()
   
      try:
         Macros = ExcelApp.Workbooks.Open(Filename=str(macro_path))
         Macros.Windows(1).Visible = False
         ExcelApp.Run("DataForgeMACROtmp.xlsm!FILEMERGE")
         Blank.Activate() 
         Macros.Close(SaveChanges=False)
      except:
         messagebox.showerror('Error', 'Macro Template Missing!')
         return
   except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

def is_excel_file_open(FileSelection):
   file_select = str(FileSelection)
   target_name = os.path.basename(file_select).lower()

   try:
      excel = win32com.client.GetActiveObject("Excel.Application")
   except Exception:
      return False

   for wb in excel.Workbooks:
      wb_fullname = str(wb.FullName)

        
      if wb_fullname.lower().startswith("http"):
         parsed = urlparse(wb_fullname)
         wb_name = os.path.basename(parsed.path).lower()
      else:
         wb_name = os.path.basename(os.path.normpath(wb_fullname)).lower()

        # Debug prints (optional)
      # print("OpenFile:", wb_fullname)
      # print("Target  :", file_select)
      # print("Compare :", wb_name, "==", target_name)

      if wb_name == target_name:
         return True

   return False
  
def on_combobox_change(event):
   # global selected_option
   selected_option = Letsgo.get()
   if selected_option == "-Choose Your Process-":
      Pselector.selection_clear()
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      cbutton.config(state="disabled")
      Location.config(text="")
      Status.config(text="")
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.clear()

   if selected_option == "AutoDraw - USDA NIFA":
      Pselector.selection_clear()
      sbutton.config(command=Autodraw)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      cbutton.config(state="disabled")
      Location.config(text="")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xlsx"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "AutoDraw - NIH":
      Pselector.selection_clear()
      sbutton.config(command=Autodraw3)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      cbutton.config(state="disabled")
      Location.config(text="")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xlsx"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "AutoDraw - NSF":
      Pselector.selection_clear()
      sbutton.config(command=Autodraw2)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      cbutton.config(state="disabled")
      Location.config(text="")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xls"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "WireTap":
      Pselector.selection_clear()
      cbutton.config(state="normal")
      sbutton.config(command=pdf_scan)
      progress_var.set(0)
      OName.config(state="normal")
      OName.delete(0, tk.END)
      SaveA.config(foreground="yellow2")
      CurrentL.config(foreground="yellow2")
      Location.config(text=dflon)
      file_types.clear()
      file_types.append(("PDF Files", "*.pdf"))
      Export.grid(row=6, column=3, columnspan=6)
      cbutton.grid(row=6, column=5, padx=10, pady=10, sticky="ne")
      OName.config(state="normal")
      OName.insert(0, Namepush)
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "PDFReader":
      Pselector.selection_clear()
      sbutton.config(command=PDFREADER)
      cbutton.config(state="normal")
      progress_var.set(0)
      OName.config(state="normal")
      OName.delete(0, tk.END)
      SaveA.config(foreground="yellow2")
      CurrentL.config(foreground="yellow2")
      Location.config(text=dflon)
      file_types.clear()
      file_types.append(("PDF Files", "*.pdf"))
      Export.grid(row=6, column=3, columnspan=6)
      cbutton.grid(row=6, column=5, padx=10, pady=10, sticky="ne")
      OName.config(state="normal")
      OName.insert(0, Namepush)
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "RetroActivity":
      Pselector.selection_clear()
      sbutton.config(command=RETROACTIVITIES)
      cbutton.config(state="normal")
      progress_var.set(0)
      OName.config(state="normal")
      OName.delete(0, tk.END)
      SaveA.config(foreground="yellow2")
      CurrentL.config(foreground="yellow2")
      Location.config(text=dflon)
      file_types.clear()
      file_types.append(("PDF Files", "*.pdf"))
      Export.grid(row=6, column=3, columnspan=6)
      cbutton.grid(row=6, column=5, padx=10, pady=10, sticky="ne")
      OName.config(state="normal")
      OName.insert(0, Namepush)
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "AZAPPER":
      Pselector.selection_clear()
      sbutton.config(command=AZAPPER)
      cbutton.config(state="normal")
      progress_var.set(0)
      OName.config(state="normal")
      OName.delete(0, tk.END)
      SaveA.config(foreground="yellow2")
      CurrentL.config(foreground="yellow2")
      Location.config(text=dflon)
      file_types.clear()
      file_types.append(("PDF Files", "*.pdf"))
      Export.grid(row=6, column=3, columnspan=6)
      cbutton.grid(row=6, column=5, padx=10, pady=10, sticky="ne")
      OName.config(state="normal")
      OName.insert(0, Namepush)
      Status.config(text="READY")
      Status.config(foreground="medium spring green")


   if selected_option == "Invictus":
      Pselector.selection_clear()
      sbutton.config(command=Invictus)
      cbutton.config(state="normal")
      progress_var.set(0)
      OName.config(state="normal")
      OName.delete(0, tk.END)
      SaveA.config(foreground="yellow2")
      CurrentL.config(foreground="yellow2")
      Location.config(text=dflon)
      file_types.clear()
      file_types.append(("PDF Files", "*.pdf"))
      Export.grid(row=6, column=3)
      # cbutton.grid_remove()
      cbutton.grid(row=6, column=5, padx=10, pady=10, sticky="ne")
      OName.config(state="normal")
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "QuickFormat":
      cbutton.config(state="disabled")
      Pselector.selection_clear()
      sbutton.config(command=QuickFormat)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      Location.config(text="")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xls"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")
      # OName.config(text="")
   if selected_option == "ASAP Drawsheet Formatter":
      cbutton.config(state="disabled")
      Pselector.selection_clear()
      sbutton.config(command=DSFormat)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      Location.config(text="")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xlsx"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")

   if selected_option == "The Number":
      cbutton.config(state="disabled")
      Pselector.selection_clear()
      sbutton.config(command=THENUMBER)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      Location.config(text="")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xlsx"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")   

   elif selected_option == "P4: Payroller":
      cbutton.config(state="disabled")
      Pselector.selection_clear()
      sbutton.config(command=Payroll)
      progress_var.set(0)
      OName.delete(0, tk.END)
      OName.config(state="disabled")
      SaveA.config(foreground="grey")
      CurrentL.config(foreground="grey")
      file_types.clear()
      Export.grid_remove()
      cbutton.grid_remove()
      file_types.append(("Excel Files", "*.xlsx"))
      Status.config(text="READY")
      Status.config(foreground="medium spring green")


Wid_frame = ttk.LabelFrame(frame, text="DATAFORGE v2.2")
Wid_frame.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

calculator = ImageTk.PhotoImage(Image.open("assets/calcicon2.png").resize((25,25)))
clipboard = ImageTk.PhotoImage(Image.open("assets/clipboard.png").resize((25,25)))
snipsnip = ImageTk.PhotoImage(Image.open("assets/scisors.png").resize((25,25)))
openfolder = ImageTk.PhotoImage(Image.open("assets/openfolder.png").resize((25,25)))
playicon = ImageTk.PhotoImage(Image.open("assets/PlayIcon.png").resize((25,25)))
quik = ImageTk.PhotoImage(Image.open("assets/quik.png").resize((25,25)))
xpand = ImageTk.PhotoImage(Image.open("assets/xpand.png").resize((25,25)))
reconitems = ImageTk.PhotoImage(Image.open("assets/ReconIcon.png").resize((35,35)))
invsumicon = ImageTk.PhotoImage(Image.open("assets/Invsumicon.png").resize((35,35)))
pmricon = ImageTk.PhotoImage(Image.open("assets/pmricon.png").resize((35,35)))
helpme = ImageTk.PhotoImage(Image.open("assets/questionicon.png").resize((25,25)))
detailarrow = ImageTk.PhotoImage(Image.open("assets/downarrow.png").resize((25,25)))
exiticon = ImageTk.PhotoImage(Image.open("assets/exiticon.png").resize((30,25)))
img = ImageTk.PhotoImage(Image.open("assets/NCSULOGO.png"))
Leftarrow = ImageTk.PhotoImage(Image.open("assets/larrow.png").resize((25,25)))
rightarrow = ImageTk.PhotoImage(Image.open("assets/rarrow.png").resize((25,25)))
selectarrow = ImageTk.PhotoImage(Image.open("assets/select.png").resize((25,25)))
Wolfpic = ImageTk.PhotoImage(Image.open("assets/NCSULOGO3.png").resize((100,100)))
moneybag = ImageTk.PhotoImage(Image.open("assets/MoneyBag.png").resize((25,25)))
importicon = ImageTk.PhotoImage(Image.open("assets/importicon.png").resize((25,25)))
folder_icon = ImageTk.PhotoImage(Image.open("assets/RedFolder2.png").resize((16, 16)))
quil_icon = ImageTk.PhotoImage(Image.open("assets/Quilicon.png").resize((16, 16)))
inspect_icon = ImageTk.PhotoImage(Image.open("assets/Inspect.png").resize((25,25)))
generate_icon = ImageTk.PhotoImage(Image.open("assets/Generate.png").resize((30,30)))
default_icon = ImageTk.PhotoImage(Image.open("assets/DefualtFile.png").resize((16, 16)))
stapler = ImageTk.PhotoImage(Image.open("assets/stapler.png").resize((25,25)))
file_icons = {
    'xlsx': ImageTk.PhotoImage(Image.open("assets/xlsfile.png").resize((16, 16))),
    'xlsm': ImageTk.PhotoImage(Image.open("assets/xlsfile.png").resize((16, 16))),
    'csv': ImageTk.PhotoImage(Image.open("assets/xlsfile.png").resize((16, 16))),
    'xls': ImageTk.PhotoImage(Image.open("assets/xlsfile.png").resize((16, 16))),
    'pdf': ImageTk.PhotoImage(Image.open("assets/pdficon.png").resize((16, 16))),
    'zip': ImageTk.PhotoImage(Image.open("assets/ZippedFolder.png").resize((16, 16))),
    'mp3': ImageTk.PhotoImage(Image.open("assets/mp3icon.png").resize((16, 16))),
    'wav': ImageTk.PhotoImage(Image.open("assets/mp3icon.png").resize((16,16))),
    'png': ImageTk.PhotoImage(Image.open("assets/pngicon.png").resize((16, 16))),
    'jpeg': ImageTk.PhotoImage(Image.open("assets/jpgicon.png").resize((16, 16))),
    'jpg': ImageTk.PhotoImage(Image.open("assets/jpgicon.png").resize((16, 16))),
    'docx': ImageTk.PhotoImage(Image.open("assets/wordicon.png").resize((16, 16))) }

panel = Label(Wid_frame, image = img)
panel.grid(row=1, column=1, padx=10, pady=10, sticky="sew")
panel.focus()

def details():
   if not Dframe.winfo_ismapped():
      Dframe.grid(row=4, column=2, columnspan=5, padx=5, pady=5)
   else:
      Dframe.grid_remove()

def open_clipboard(): 
   pyautogui.keyDown('win')
   pyautogui.press('v')
   pyautogui.keyUp('win')

def expand():
   pyautogui.keyDown('win')
   pyautogui.press('tab')
   pyautogui.keyUp('win')

def launch_snipping_tool():
   try:
      subprocess.run('snippingtool.exe', check=True)
   except FileNotFoundError:
      print("Snipping Tool not found!")

def launch_calculator():
   try:
      subprocess.run('calc.exe', check=True)
   except FileNotFoundError:
      print("Calculator app not found!")

def clear():
   Budgetinfo.config(state="normal")
   Budgetinfo.delete(0, tk.END)
   # Budgetinfo.config(state="readonly")
   WRSCashB.config(state="normal")
   WRSCashB.delete(0, tk.END)
   # WRSCashB.config(state="readonly")
   WRSRecivables.config(state="normal")
   WRSRecivables.delete(0, tk.END)
   # WRSRecivables.config(state="readonly")
   PTDXpenses.config(state="normal")
   PTDXpenses.delete(0, tk.END)
   WRSTotalExpenses.config(state="normal")
   WRSTotalExpenses.delete(0, tk.END)
   WRSIndirectcost.config(state="normal")
   WRSIndirectcost.delete(0, tk.END)
   WRSIDbudget.config(state="normal")
   WRSIDbudget.delete(0, tk.END)
   WRSDbudget.config(state="normal")
   WRSDbudget.delete(0, tk.END)
   PTDExpensesFR.config(state="normal")
   PTDExpensesFR.delete(0, tk.END)
   PIname.config(state="normal")
   PIname.delete(0, tk.END)

Toolbar = ttk.Frame(Wid_frame)
Toolbar.grid(row=1, column=2, columnspan=5, padx=10, pady=10, sticky="w")

Sniptool = ttk.Button(Toolbar, text="Snip Tool", image=snipsnip, compound="top", command=start_snip_thread)
Sniptool.grid(row=1, column=0, sticky="e")

Additup = ttk.Button(Toolbar, text="Calculator", image=calculator, compound="top", command=launch_calculator)
Additup.grid(row=1, column=2, sticky="e")

Clipboardtool = ttk.Button(Toolbar, text="Clipboard", image=clipboard, compound="top", command=open_clipboard)
Clipboardtool.grid(row=1, column=3, sticky="e")

TaskViewer = ttk.Button(Toolbar, text="Expand", image=xpand, compound="top", command=expand)
TaskViewer.grid(row=1, column=4, sticky="e")

Together = ttk.Button(Toolbar, text="Stapler", image=stapler, compound="top", command=Consolidate)
Together.grid(row=1, column=5, sticky="e")

Help = ttk.Button(Toolbar, text="Help", compound="top", image=helpme, command=GetHelp)
Help.grid(row=1, column=6, sticky="e")

FileType = ttk.LabelFrame(Wid_frame, text="File Select")
FileType.grid(row=2, column=2, columnspan=4, padx=10, pady=5, sticky="nsew")

FileName1 = ttk.Label(FileType, text=file_path)
FileName1.grid(row=3, column=3, padx=3, sticky="sw")

DetailBar = ttk.Frame(Wid_frame)
# DetailBar.grid(row=3, column=2, columnspan=2, padx=10, pady=5, sticky="nsew")

DocumentType = ttk.Label(DetailBar, image="")
DocumentType.grid(row=3, column=2)

DocStatus = tk.Label(DetailBar, text="")
DocStatus.grid(row=3, column=4, padx=5, pady=5, sticky="ew")

ReportType = tk.Label(DetailBar, text="", fg="black")
ReportType.grid(row=3, column=3, padx=5, pady=5)

ProjectID = tk.Label(DetailBar, text="", fg="yellow")
ProjectID.grid(row=3, column=5, padx=5, pady=5, sticky="ew")

RRv = IntVar()
RRv.set(0)

Checkrb = IntVar()
Checkrb.set(0)

SF425v = IntVar()
SF425v.set(0)

SF270v = IntVar()
SF270v.set(0)

DD882v = IntVar()
DD882v.set(0)

SRbv = IntVar()
SRbv.set(0)

Ivs = IntVar()
Ivs.set(0)

def open_import_menu():
   global import_menu
   import_menu = tk.Toplevel(root)
   import_menu.title("Import Menu")

   InternalOptions = ['Reconciling Items','PMR Worksheet', 'Refund Request', 'Cost Share Memo']
   # ExtenalOptions = ['SF-425', 'SF-270', 'Generic Financial Report', 'SF-270 Worksheet']

   # selected_option = tk.StringVar(value=InternalOptions[0])


   ok_button = ttk.Button(import_menu, text="OK", command=close_import_menu)
   ok_button.grid(row=10, column=1, pady=10, sticky="ew")

   InternalForms = ttk.LabelFrame(import_menu, text="Internal Forms")
   # InternalForms.grid(row=1, column=1, padx=5, pady=5, sticky="w")

   ReconRadio = ttk.Checkbutton(InternalForms, text="Reconciling Items Worksheet", image=reconitems, variable=RRv, compound="left")
   ReconRadio.grid(row=1, column=1, padx=5, pady=5, sticky="w")

   InvoiceSummary = ttk.Checkbutton(InternalForms, text="Invoice Summary", image=invsumicon, compound="left", variable=Ivs)
   InvoiceSummary.grid(row=7, column=1, padx=5, pady=5, sticky="w")

   CustomerID = ttk.Checkbutton(InternalForms, text="Customer ID")
   CustomerID.grid(row=8, column=1, padx=5, pady=5, sticky="w")

   Loadb2= tk.Button(InternalForms, text="Load  ", image=file_icons['xlsx'], compound="right")
   # Loadb2.grid(row=8, column=2, padx=10, pady=10, sticky="w")

   FileSelect = ttk.LabelFrame(InternalForms)
   # FileSelect.grid(row=8, column=1)

   IvSummary = ttk.Label(FileSelect, text="No File Selected")
   # IvSummary.grid(row=1, column=1)

   SRebudget = ttk.Checkbutton(InternalForms, text="PMR Summary", variable=SRbv, image=pmricon, compound="left")
   SRebudget.grid(row=5, column=1, padx=5, pady=5, sticky="w")

   Loadb3= tk.Button(InternalForms, text="Load  ", image=file_icons['xlsx'], compound="right")
   # Loadb3.grid(row=6, column=1, padx=10, pady=10, sticky="w")
   import_menu.resizable(False, False)
   import_menu.transient(root)
   import_menu.grab_set()

def close_import_menu():
   if RRv.get() == 1:
      ReconLight.grid(row=6, column=5, padx=5, pady=5, sticky="w")
   if RRv.get() == 0:
      ReconLight.grid_remove()
   if SRbv.get() == 1:
      PMRLight.grid(row=6, column=6, padx=5, pady=5, sticky="w")
   if SRbv.get() == 0:
      PMRLight.grid_remove()
   import_menu.destroy()    

def open_sub_menu():
    # Create a new window for the sub-menu
   global sub_menu
   sub_menu = tk.Toplevel(root)
   sub_menu.title("Generate")

   Wolfy = ttk.Label(sub_menu, image=Wolfpic)
   # Wolfy.grid(row=0, column=1, rowspan=2, padx=5, pady=10, sticky="ns")

   sarrow = ttk.Label(sub_menu, image=selectarrow)
   sarrow.grid(row=1, column=6, padx=5, pady=5, sticky="nsew")

   Leftbutton = tk.Button(sub_menu, image=Leftarrow)
   # Leftbutton.grid(row=0, column=2, sticky="w")

   Rightbutton = tk.Button(sub_menu, image=rightarrow)
   # Rightbutton.grid(row=0, column=6, sticky="w")

   Profiles = "Select A Profile", "New Award Review", "Financial Reporting", "A/R", "Invoicing", "Closeout", "Federal Draw", "Custom Profile"

   ProfileSelect = ttk.Combobox(sub_menu, values=Profiles, state="readonly")
   # ProfileSelect.grid(row=0, column=3, columnspan=2, sticky="nsew")
   ProfileSelect.current(0)

   StatusFrame = ttk.LabelFrame(sub_menu, text="STATUS:")
   StatusFrame.grid(row=10, column=4, padx=5, pady=5, sticky="nsew")

   global Status
   Status = ttk.Label(StatusFrame, text="")
   Status.grid(row=1, column=1, padx=5, pady=5, sticky="nsew" )

   global Pselector

   Pselector = ttk.Combobox(sub_menu, values=Reports, state="readonly", textvariable=Letsgo)
   Pselector.grid(row=1, column=3, columnspan=3, padx=10, pady=10, sticky="nsew")
   Pselector.bind("<<ComboboxSelected>>", on_combobox_change)
   if file_extension in ['pdf']:
      Pselector.config(values=PDFReports)
   elif file_extension in ['xlsx'] or ['xls']:
      Pselector.config(values=XLXSReports)
   else:
      Pselector.config(values=Reports)
   Pselector.current(0)

   global sbutton
   sbutton = ttk.Button(sub_menu, text="Start  ", command=start_process_thread, image=playicon, compound="right")
   sbutton.grid(row=10, column=3, pady=10, padx=10, sticky="w")

   # FileName2 = ttk.Label(Slot2, text=file_path2)
   # FileName2.grid(row=4, column=2, padx=20, pady=10)
   
   FileType2 = ttk.LabelFrame(sub_menu, text="File Select 2")
   # FileType2.grid(row=2, column=3, columnspan=4, padx=10, pady=5, sticky="nsew")

   global FileName2
   FileName2 = ttk.Label(FileType2, text=file_path2)
   # FileName2.grid(row=3, column=4, padx=3, sticky="sw")

   Loadb2= tk.Button(sub_menu, text="Load  ", command=Load2, image=openfolder, compound="right")
   # Loadb2.grid(row=2, column=2, padx=10, pady=10, sticky="se")

   ExternalForms = ttk.LabelFrame(sub_menu, text="External Forms")
   # ExternalForms.grid(row=3, column=3, columnspan=2, padx=5, pady=5, sticky="ew")

   InternalForms = ttk.LabelFrame(sub_menu, text="Internal Forms")
   # InternalForms.grid(row=2, column=3, padx=5, columnspan=2, pady=5, sticky="ew")

   SF425Radio = ttk.Checkbutton(ExternalForms, text="Federal SF-425", variable=SF425v)
   SF425Radio.grid(row=1, column=1, padx=5, pady=5, sticky="w")

   SF270Radio = ttk.Checkbutton(ExternalForms, text="Federal SF-270", variable=SF270v)
   SF270Radio.grid(row=2, column=1, padx=5, pady=5, sticky="w")

   DD882Radio = ttk.Checkbutton(ExternalForms, text="Federal DD-882", variable=DD882v)
   DD882Radio.grid(row=3, column=1, padx=5, pady=5, sticky="w")

   CheckRequest = ttk.Checkbutton(InternalForms, text="Check Refund Request", variable=RRv)
   CheckRequest.grid(row=2, column=1, padx=5, pady=5, sticky="w")

   ASAPRequest = ttk.Checkbutton(InternalForms, text="ASAP Refund Request", variable=RRv)
   ASAPRequest.grid(row=3, column=1, padx=5, pady=5, sticky="w")

   SRebudget = ttk.Checkbutton(InternalForms, text="Rebudget Request Worksheet", variable=SRbv)
   SRebudget.grid(row=4, column=1, padx=5, pady=5, sticky="w")

   CostshareMemo = ttk.Checkbutton(InternalForms, text="Cost Share Memo")
   CostshareMemo.grid(row=4, column=1, padx=5, pady=5, sticky="w")

   Exitbutton = ttk.Button(sub_menu, text="Exit  ", image=exiticon, command=close_sub_menu, compound="right")
   Exitbutton.grid(row=10, column=5, padx=5, pady=5, sticky="e")

   global Export
   Export = ttk.Frame(sub_menu)
   OutputName = ttk.Frame(Export)
   OutputName.grid(row=4, column=0, columnspan=4, sticky="ew")

   global SaveA
   SaveA = ttk.Label(OutputName, text="Output Name:", image=quil_icon, compound="right", foreground="grey")
   SaveA.grid(row=2, column=2, padx=10, pady=10, sticky="w")

   global CurrentL
   CurrentL = ttk.Label(OutputName, text="Save Location:", image=folder_icon, compound="right", foreground="grey")
   CurrentL.grid(row=0, column=2, padx=10, pady=10, sticky="new")

   global OName
   OName = ttk.Entry(OutputName, textvariable=default_name)
   OName.grid(row=2, column=3, columnspan=2, padx=10, pady=10, sticky="ew")

   global Location
   Location = ttk.Label(OutputName, font="bold")
   Location.grid(row=0, column=3, padx=15, pady=10, sticky="new")

   global cbutton
   cbutton = tk.Button(sub_menu, text="Change", command=get_file_path, state="disabled")
   
   sub_menu.resizable(False, False)
   sub_menu.transient(root) 
   sub_menu.grab_set()

def close_sub_menu():
   sub_menu.destroy()

def QuickFormat():
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
      return
   if file_extension == 'xlsx':
      messagebox.showerror('File Already Formatted!', 'Please Choose a Valid File')
      return
   if file_extension != 'xls':
      messagebox.showerror('File Error', 'Please Choose a Valid File')
      return
   else:
      # clear()
      try:
         DocStatus.config(text="")
         ReportType.config(text="RUNNING")
         ReportType.config(fg="VioletRed1")

         ExcelApp = win32com.client.Dispatch("Excel.Application")

         # ExcelApp.visible = False
         # FS2 = Path(FileSelection)
         try:
            QuikFormat = ExcelApp.Workbooks.Open(SelectedFile)
         except:
            messagebox.showerror('File Already Formatted!', 'Please Choose a Valid File')
            return
         progress_var.set(0)
         for sheet_index in range(1, QuikFormat.Sheets.Count + 1):
            progress_var.set(sheet_index)
            root.update_idletasks()
            time.sleep(0.1)
            if sheet_index != "Sheet1":
               WRSsheet = QuikFormat.Sheets(sheet_index)
               TRvalidation = WRSsheet.Cells(24, "A").Value
               clean_value = TRvalidation.replace("\xa0", " ").strip() 
               print(clean_value)
               if clean_value != "40000-49999":
                  WRSsheet.Rows(24).Insert()
              
               WRSsheet.Cells.UnMerge()
               WRSsheet.Rows(19).Delete()
               WRSsheet.Rows(10).Delete()
               WRSsheet.Rows(29).Delete()
               WRSsheet.Rows(39).Delete()
               WRSsheet.Rows(50).Delete()
               WRSsheet.Rows(50).Delete()
               WRSsheet.Rows(50).Delete()
               WRSsheet.Rows(50).Delete()
               WRSsheet.Columns(7).Insert()
               WRSsheet.Columns(7).Insert()
               WRSsheet.Columns(7).Insert()
               WRSsheet.Cells(16, "F").Value = "Reconciling Items"
               WRSsheet.Cells(16, "G").Value = "Final Expenses"
               WRSsheet.Cells(16, "H").Value = "Invoice Cummulative"
               WRSsheet.Cells(16, "I").Value = "Final Invoice"
               BudgetTotalFormula = "=SUM(C23:C46)"
               CurrentBudgetTotal = "=C47+C48"
               CMTFormula = "=D47+D48"
               PTDFormula = "=E47+E48"
               CurrentMonthTotal = "=SUM(D23:D46)"
               PTDActivityFormula = "=SUM(E23:E46)"
               ReconcilingTotal = "=SUM(F23:F46)"
               FinalExpensesFormula ="=SUM(G23:G46)"
               BudgetBTFormula = "=SUM(J23:J46)"
               BudgetBIDCFormula = "=C48-G48"
               BudgetBalanceF = "=C23-G23"
               FinalXpense = "=E23+F23"
               VcummFormula = "=SUM(H23:H46)"
               FinalInvoiceFormula = "=SUM(I23:I46)"
               FinalIDC = "=E48+F48"
               FinalTotal = "=G47+G48"
               InvoiceFinal = "=G23-H23"
               ReconSum = "=F47+F48"
               FINVIDC = "=G48-H48"
               FAratecheck = "=C51*$B$52"

               DataRange = WRSsheet.Range("C17:J52")

               WRSsheet.Cells(47, "C").Formula = BudgetTotalFormula
               WRSsheet.Cells(49, "C").Formula = CurrentBudgetTotal
               WRSsheet.Cells(47, "D").Formula = CurrentMonthTotal
               WRSsheet.Cells(49, "D").Formula = CMTFormula
               WRSsheet.Cells(47, "E").Formula = PTDActivityFormula
               WRSsheet.Cells(49, "E").Formula = PTDFormula
               WRSsheet.Cells(47, "F").Formula = ReconcilingTotal
               WRSsheet.Cells(47, "G").Formula = FinalExpensesFormula
               WRSsheet.Cells(47, "H").Formula = VcummFormula
               WRSsheet.Cells(47, "I").Formula = FinalInvoiceFormula
               WRSsheet.Cells(48, "G").Formula = FinalIDC
               WRSsheet.Cells(49, "G").Formula = FinalTotal
               WRSsheet.Cells(48, "I").Formula = FINVIDC
               WRSsheet.Cells(47, "J").Formula = BudgetBTFormula
               WRSsheet.Cells(48, "J").Formula = BudgetBIDCFormula
               
               # print(EHRALabel)
               WRSsheet.Cells(23, "G").Formula = FinalXpense
               FinalRange = WRSsheet.Range("G23:G46")
               WRSsheet.Cells(23, "G").AutoFill(FinalRange)

               DirectTotal = WRSsheet.Range("G49:J49")
               WRSsheet.Cells(49, "G").AutoFill(DirectTotal)

               WRSsheet.Cells(23, "I").Formula = InvoiceFinal
               FINVrange = WRSsheet.Range("I23:I46")
               WRSsheet.Cells(23, "I").AutoFill(FINVrange)

               WRSsheet.Cells(23, "J").Formula = BudgetBalanceF
               BBrange = WRSsheet.Range("J23:J46")
               WRSsheet.Cells(23, "J").AutoFill(BBrange)

               WRSsheet.Cells(23, "F").Value = 0
               Encumb = WRSsheet.Range("F23:F46")
               WRSsheet.Cells(23, "F").AutoFill(Encumb)
               WRSsheet.Cells(49, "F").Formula = ReconSum

               IDCrate = WRSsheet.Range("F10").Value
               value = str(IDCrate)

               if "%" in value:
                  value = value.replace("%", "")

               if value == "None":
                  pass
               else:
                  WRSsheet.Range("F10").Value = value
                  IDCrate2 = WRSsheet.Range("F10").Value
                  IDC = re.findall(r'\b\d+(?:\.\d{1,3})?\b', str(IDCrate2))
                  Idcnumber = float(IDC[0])
                  BASErate = re.findall(r'[A-Z]', str(IDCrate2))
                  Rate = "".join(BASErate)

                  print(IDC)
                  IDCicky = WRSsheet.Range("B52")
                  WRSsheet.Range("B51").Value = Rate
                  IDCicky.Value = Idcnumber / 100
                  IDCicky.NumberFormat = "0%"
                  # c42 = WRSsheet.Range("C42").Value
                  # c47 = WRSsheet.Range("C47").Value
                  # print("C42 =", repr(c42), type(c42))
                  # print("C47 =", repr(c47), type(c47))
                  if WRSsheet.Range("B51").Value == "MTDC":
                     if float(WRSsheet.Range("C42").Value) == float(WRSsheet.Range("C47").Value):
                        if WRSsheet.Range("C42").Value < 25000:
                           DirectCost = WRSsheet.Range("C42").Value
                           WRSsheet.Range("C51").Value = DirectCost
                           WRSsheet.Range("E51").Value = WRSsheet.Range("E42")
                           WRSsheet.Range("G51").Value = WRSsheet.Range("G42")
                           WRSsheet.Range("I51").Value = WRSsheet.Range("I51")
                        else:
                           DirectCost = 25000
                           WRSsheet.Range("C51").Value = DirectCost
                           WRSsheet.Range("E51").Value = DirectCost
                           WRSsheet.Range("G51").Value = DirectCost
                           WRSsheet.Range("I51").Value = DirectCost
                     else:
                        FAcheckFormula = "=C47-C39-C40-C41-C37-C38-C42"
                        WRSsheet.Range("C51").Value = FAcheckFormula
                        FRange = WRSsheet.Range("C51:J51")
                        WRSsheet.Range("C51").AutoFill(FRange)
                     WRSsheet.Range("C52").Value = FAratecheck
                     IDCrange = WRSsheet.Range("C52:J52")
                     WRSsheet.Range("C52").AutoFill(IDCrange)


                  if WRSsheet.Range("B51").Value == "TDC":
                     if float(WRSsheet.Range("C42").Value) == float(WRSsheet.Range("C47").Value):
                        if WRSsheet.Range("C42").Value < 25000:
                           DirectCost = WRSsheet.Range("C42").Value
                           WRSsheet.Range("C51").Value = DirectCost
                           WRSsheet.Range("E51").Value = WRSsheet.Range("E42")
                           WRSsheet.Range("G51").Value = WRSsheet.Range("G42")
                           WRSsheet.Range("I51").Value = WRSsheet.Range("I51")
                        else: 
                           DirectCost = 25000
                           WRSsheet.Range("C51").Value = DirectCost
                           WRSsheet.Range("E51").Value = DirectCost
                           WRSsheet.Range("G51").Value = DirectCost
                           WRSsheet.Range("I51").Value = DirectCost
                     else:
                        DirectCost = "=C47"
                     WRSsheet.Range("C51").Value = DirectCost
                     FRange = WRSsheet.Range("C51:J51")
                     WRSsheet.Range("C51").AutoFill(FRange)
                     WRSsheet.Range("C52").Value = FAratecheck
                     IDCrange = WRSsheet.Range("C52:J52")
                     WRSsheet.Range("C52").AutoFill(IDCrange)

               
                  DataRange.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
                  DataRange.HorizontalAlignment = c.xlLeft
                  
         if file_extension == 'xls': 
            xlsx_file = str(SelectedFile) + "x"
            QuikFormat.SaveAs(xlsx_file, FileFormat=51)
            QuikFormat.Close()
            pythoncom.CoUninitialize()
            os.remove(SelectedFile)
            global FS2
            FS2 = Path(xlsx_file)
            FileName1.config(text=FS2.name)
            df4 = pd.ExcelFile(xlsx_file)
            sheetnames1 = df4.sheet_names
            WRSsearch = re.compile(r'WRS_Report_P1_')
            WRStag1 = [name for name in sheetnames1 if WRSsearch.search(name)]
            # print(workbook)
            try:
               if WRStag1:
                  df = pd.read_excel(xlsx_file)
                  DetailButton.grid()
                  ReportType.config(text="P1:")
                  ReportType.config(fg="yellow")
                  DetailButton.config(state="normal")
                  ProjectIDregex = r'[50]\d{5}'
                  matche = re.search(ProjectIDregex, ShortName)
                  if matche:
                     ProjectID.config(text=matche.group())
                  SponsorName = df.iloc[12,1]
                  AsofDate = df.iloc[0,0]
                  Pinvestigator = df.iloc[4,3]
                  DocStatus.config(text=SponsorName)
                  DocStatus.config(fg="white")
                  OUCpanda = df.iloc[4,5]
                  Alltime = df.iloc[7,3]
                  FArateinto = df.iloc[8,5]
                  Awardtotal = df.iloc[47,2]
                  TotalIDC = df.iloc[46,2]
                  TotalDC = df.iloc[45,2]
                  WRSKASH = df.iloc[15,4]
                  OtherReceiveables = df.iloc[16,4]
                  PTDTotal = df.iloc[47,4]
                  PTDIDC = df.iloc[46,4]
                  PTDDC = df.iloc[45,4]
                  clear()
                  # Placeholder.config(text=AsofDate)
                  P1Summary.config(text=f"Summary - {AsofDate}")
                  P1Summary.config(fg="orange")
                  PIname.insert(0, str(Pinvestigator))
                  PIname.config(state="readonly")
                  POPInfo.config(text=Alltime)
                  OUCcode.insert(0,str(OUCpanda))
                  OUCcode.config(state="readonly")
                  FaInfo.config(text=FArateinto)
                  FaInfo.config(fg="yellow")
                  Budgetinfo.insert(0,str('{:,.2f}'.format(Awardtotal)))
                  Budgetinfo.config(state="readonly")
                  WRSCashB.insert(0,str('{:,.2f}'.format(WRSKASH)))
                  WRSCashB.config(state="readonly")
                  WRSRecivables.insert(0,str('{:,.2f}'.format(OtherReceiveables)))
                  WRSRecivables.config(state="readonly")
                  PTDXpenses.insert(0,str('{:,.2f}'.format(PTDTotal)))
                  PTDXpenses.config(state="readonly")
                  WRSTotalExpenses.insert(0,str('{:,.2f}'.format(PTDDC)))
                  WRSTotalExpenses.config(state="readonly")
                  WRSIndirectcost.insert(0,str('{:,.2f}'.format(PTDIDC)))
                  WRSIndirectcost.config(state="readonly")
                  # WRSIDbudget.delete(0, tk.END)
                  WRSIDbudget.insert(0,str('{:,.2f}'.format(TotalIDC)))
                  WRSIDbudget.config(state="readonly")
                  WRSDbudget.insert(0,str('{:,.2f}'.format(TotalDC)))
                  WRSDbudget.config(state="readonly")
                  if WRSKASH < 0:
                     CashB.config(fg="orange")
                  if WRSKASH > 0:
                     CashB.config(fg="red")
                  if WRSKASH == 0:
                     CashB.config(fg="SeaGreen1")   
                  if abs(OtherReceiveables) == abs(Awardtotal):
                     Receivables.config(fg="SeaGreen1")
                  if OtherReceiveables == 0:
                     Receivables.config(fg="red")
                  if OtherReceiveables < 0:
                     Receivables.config(fg="orange")
                  print(abs(WRSKASH))
                  print(abs(Awardtotal))
                  if PTDTotal > Awardtotal:
                     Totalcost.config(fg="red")
                  if PTDTotal < Awardtotal:
                     Totalcost.config(fg="sky blue")
                  if PTDDC > TotalDC:
                     DirectExpenses.config(fg="red")
                  if PTDDC < TotalDC:
                     DirectExpenses.config(fg="sky blue")
                  if PTDIDC > TotalIDC:
                     Indirectcost.config(fg="red")
                  if PTDIDC < TotalIDC:
                     Indirectcost.config(fg="sky blue")
                  if PTDIDC == TotalIDC:
                     Indirectcost.config(fg="sky blue")

                  PTDExpensesFR.insert(0,str('{:,.2f}'.format(PTDTotal)))
                  PTDExpensesFR.config(state="readonly")
            except Exception as e:
               print(f"An error occurred: {e}")


         if file_extension == 'xlsx':
            QuikFormat.Save()
            QuikFormat.Close()


               # xlsx_file2 = str(SelectedFile)
               # QuikFormat.Save()
               # QuikFormat.Close()
            # QuikFormat.Close()
         progress_var.set(100)
         Plusbutton.grid(row=15, column=1, columnspan=6, pady=5, padx=5, sticky="nsew")
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

QuickFormatButton = ttk.Button(Wid_frame, image=quik, command=QuickFormat)
QuickFormatButton.grid(row=2, column=5, padx=5, pady=5, sticky="se")

DetailButton = ttk.Button(Wid_frame, image=detailarrow, command=lambda:details())
DetailButton.grid(row=3, column=5, padx=5, pady=5, sticky="se")
DetailButton.grid_remove()

Dframe = ttk.Frame(Wid_frame)

P1Summary = tk.LabelFrame(Dframe, text="Summary")
P1Summary.grid(row=0, column=3, rowspan=3, columnspan=4, padx=5, pady=5)
# P1Summary.grid_remove()

# Placeholder = ttk.Label(P1Summary, text="")
# Placeholder.grid(row=1, column=2)
PIname = ttk.Entry(P1Summary)
PIname.grid(row=3, column=2, padx=5, pady=5, sticky="nsew")

OUCcode = ttk.Entry(P1Summary)
OUCcode.grid(row=3, column=4, padx=5, pady=5, sticky="nsew")

PrincInv = ttk.Label(P1Summary, text="PI:")
PrincInv.grid(row=3, column=1, padx=5, sticky="nsew")

Department = ttk.Label(P1Summary, text="Dept:")
Department.grid(row=3, column=3, padx=5, sticky="nsew")

PeriodofP = ttk.Label(P1Summary, text="Period:")
PeriodofP.grid(row=2, column=1, padx=5, pady=5, sticky="ne")

POPInfo = ttk.Label(P1Summary, text="N/A")
POPInfo.grid(row=2, column=2, padx=5, pady=5, sticky="sw")

FacilitiesRate = tk.Label(Dframe, text="F&A Rate:")
FacilitiesRate.grid(row=11, column=4, padx=5, pady=5, sticky="ne")

ImportButton = tk.Button(Dframe, text="Import", command=open_import_menu, image=importicon, bg="gray", compound="right")
# ImportButton.grid(row=6, column=4, padx=5, pady=5, sticky="nsew")

Iconframe = tk.Frame(Dframe)
Iconframe.grid(row=6, column=5, sticky="w")

ReconLight = tk.Label(Iconframe, image=reconitems)
PMRLight = tk.Label(Iconframe, image=pmricon)

DirectBudget = tk.Label(Dframe, text="Direct Budget:")
DirectBudget.grid(row=7, column=4, padx=5, pady=5, sticky="ne")

WRSDbudget = ttk.Entry(Dframe, state="normal")
WRSDbudget.grid(row=7, column=5, padx=5, pady=5, sticky="se")

IndirectBudget = tk.Label(Dframe, text="Indirect Budget:")
IndirectBudget.grid(row=8, column=4, padx=5, pady=5, sticky="ne")

WRSIDbudget = ttk.Entry(Dframe)
WRSIDbudget.grid(row=8, column=5, padx=5, pady=5, sticky="ne")

separator2 = ttk.Separator(Dframe)
separator2.grid(row=9, column=4, columnspan=2, sticky="ew")

Awardtotal = tk.Label(Dframe, text="Award Total:")
Awardtotal.grid(row=10, column=4, padx=5, pady=5, sticky="ne")

Budgetinfo = ttk.Entry(Dframe)
Budgetinfo.grid(row=10, column=5, padx=5, pady=5, sticky="ne")

FaInfo = tk.Label(Dframe, text="N/A")
FaInfo.grid(row=11, column=5, padx=5, pady=5, sticky="sw")
# Slot1 = ttk.LabelFrame(FileType, text="Slot1")
# Slot1.grid(row=2, column=1, columnspan=2, padx=20, pady=10, sticky="nsew")


Plusbutton = tk.Button(Wid_frame, text="Generate", image=generate_icon, compound="top", command=open_sub_menu, bg="gray")
# Plusbutton.grid(row=15, column=4, columnspan=3, pady=5, padx=10, sticky="ew")

ExpensesPTD = tk.Label(Dframe, text=": PTD Expenses")
# ExpensesPTD.grid(row=3, column=7, padx=5, pady=5, sticky="sw")

PTDExpensesFR = ttk.Entry(Dframe)
# PTDExpensesFR.grid(row=3, column=6, padx=5, pady=5, sticky="se")

Receivables = tk.Label(Dframe, text=": Receivables  ", image=moneybag, compound="right")
Receivables.grid(row=11, column=7, padx=5, pady=5, sticky="sw")

WRSRecivables = ttk.Entry(Dframe)
WRSRecivables.grid(row=11, column=6, padx=5, pady=5, sticky="se")

separator1 = ttk.Separator(Dframe)
separator1.grid(row=5, column=4, columnspan=4, sticky="ew")

CashB = tk.Label(Dframe, text=": Cash Balance")
CashB.grid(row=12, column=7, padx=5, pady=5, sticky="sw")

WRSCashB = ttk.Entry(Dframe)
WRSCashB.grid(row=12, column=6, padx=5, pady=5, sticky="se")

DirectExpenses = tk.Label(Dframe, text=": Direct WRS")
DirectExpenses.grid(row=7, column=7, padx=5, pady=5, sticky="sw")

WRSTotalExpenses = ttk.Entry(Dframe)
WRSTotalExpenses.grid(row=7, column=6, padx=5, pady=5, sticky="se")

Indirectcost = tk.Label(Dframe, text=": Indirect WRS")
Indirectcost.grid(row=8, column=7, padx=5, pady=5, sticky="sw")

WRSIndirectcost = ttk.Entry(Dframe)
WRSIndirectcost.grid(row=8, column=6, padx=5, pady=5, sticky="se")

separator = ttk.Separator(Dframe)
separator.grid(row=9, column=6, columnspan=2, sticky="ew")

Totalcost = tk.Label(Dframe, text=": Total WRS")
Totalcost.grid(row=10, column=7, padx=5, pady=5, sticky="sw")

PTDXpenses = ttk.Entry(Dframe)
PTDXpenses.grid(row=10, column=6, padx=5, pady=5, sticky="se")

def process_completed_handler(e):
   messagebox.showinfo('Reporting Complete!', 'Your report has been generated succesfully!')
   try:
      close_sub_menu()
   except Exception as e:
      return

# Slot2 = ttk.LabelFrame(FileType, text="Progress")
progress_var = tk.DoubleVar()
progressbar = ttk.Progressbar(Wid_frame, orient=HORIZONTAL, variable=progress_var, mode='determinate', maximum=100)
progressbar.grid(row=7, column=0, pady=5, padx=5, columnspan=8, sticky="nsew" )
# Slot2.grid(row=3, column=1, columnspan=2, padx=20, pady=10, sticky="nsew")

Bouncetype = IntVar()
Bouncetype.set(0)

Offline = ttk.Checkbutton(Wid_frame, text="Offline", variable=Bouncetype, state="disabled")
# Offline.grid(row=8, column=2, pady=5, padx=5, sticky="w")

Process = ttk.Frame(Wid_frame)
Process.grid(row=8, column=2, padx=5, pady=5, sticky="nsw")

# Pselector = ttk.Combobox(Process, values=Reports, state="readonly", textvariable=Letsgo,)
# Pselector.grid(row=1, column=1, padx=20, pady=10)
# Pselector.bind("<<ComboboxSelected>>", on_combobox_change)
# Pselector.current(0)

# Export = ttk.Frame(sub_menu)
# Export.grid(row=6, column=2)

os.chdir(defaultlo)
# SelectedFile = ttk.Label(Wid_frame, text=file_path)
# SelectedFile.grid(row=1, column=2, padx=20, pady=10)
global pathobject
pathobject = tk.StringVar()
pathobject = defaultlo

def Load():
   # global FileSelection
   global FileSelection
   FileSelection = filedialog.askopenfilename(filetypes = file_types)
   global SelectedFile
   SelectedFile = Path(FileSelection)
   # FS2 = Path(FileSelection)
   # print(SelectedFile)

   global ShortName
   ShortName = SelectedFile.name
   print(ShortName)
   FileName1.config(text=ShortName)
   global Namepush
   Namepush = os.path.splitext(ShortName)[0]
   print(Namepush)
   # if FileSelection:
        # Check for hidden temporary files
         # if file_path.startswith('~$'):
         #    return
   global file_extension
   file_extension = FileSelection.split('.')[-1]
   print(file_extension)
   
   if is_excel_file_open(FileSelection):
      ReportType.config(text="")
      # clear()
      Dframe.grid_remove()
      DetailButton.grid_remove()
      Inspector.grid_remove()
      DetailBar.grid(row=3, column=2, columnspan=2, padx=10, pady=5, sticky="nsew")
      Plusbutton.grid_remove()
      DocStatus.config(text="OPEN FILE")
      DocStatus.config(fg="red")
      DocumentType.config(image=default_icon) 
      messagebox.showerror('Selected File is Already Open' , 'Please Close Workbook to Proceed with Action!')
      
   else:
      print(f"The file '{SelectedFile}' is NOT open. You can proceed.")
      if FileSelection:     
      
         # DetailButton.grid()
         # df = pd.read_excel(FileSelection)
         
         DetailBar.grid(row=3, column=2, columnspan=2, padx=10, pady=5, sticky="nsew")
         # Inspector.grid()
         # print(sheetnames)
         if file_extension in ['docx']:
            icon = file_icons.get(file_extension, default_icon)
            DocumentType.config(image=icon)
            DocStatus.config(text="WORD")
            DocStatus.config(fg="cornflower blue")
            Dframe.grid_remove()
            DetailButton.grid_remove()
            Plusbutton.grid_remove()

         if file_extension in ['xls']:
            icon = file_icons.get(file_extension, default_icon)
            DocumentType.config(image=icon)
            Dframe.grid_remove()
            DetailButton.grid_remove()
            Plusbutton.grid_remove()
            # DetailButton.config(state="disabled")
            ReportType.config(text="")
            ProjectID.config(text="")
            DocStatus.config(text="UNFORMATTED")
            DocStatus.config(fg="red")

         if file_extension in ['pdf']:
            clear()
            Plusbutton.grid(row=15, column=1, columnspan=6, pady=5, padx=5, sticky="nsew")
            DocStatus.config(text="PDF")
            DocStatus.config(fg="tomato2")
            ProjectID.config(text="")
            ReportType.config(text="")
            icon = file_icons.get(file_extension, default_icon)
            DocumentType.config(image=icon)
            Dframe.grid_remove()
            DetailButton.grid_remove()


         if file_extension in ['xlsx']:
            # df = pd.read_excel(FileSelection)
            # global df
            clear()
            ProjectID.config(text="")
            ReportType.config(text="")
            DocStatus.config(fg="lime green")
            DocStatus.config(text="XLSX")
            Dframe.grid_remove()
            DetailButton.grid_remove()
            icon = file_icons.get(file_extension, default_icon)
            DocumentType.config(image=icon)
            Df3 = pd.ExcelFile(FileSelection)
            Plusbutton.grid(row=15, column=1, columnspan=6, pady=5, padx=5, sticky="nsew")
            try:
               LOCfinder = pd.read_excel(FileSelection)
               LOCvalid = LOCfinder.iloc[1,0]
               LOCdate = LOCfinder.iloc[2,0]
               LOCSponsor = LOCfinder.iloc[6,1]
               LOCAgency = LOCfinder.iloc[4,1]
               if LOCvalid == "Monthly Letter of Credit[LOC] Report":
                  clear()
                  # DetailButton.grid()
                  Dframe.grid_remove()
                  ReportType.config(text=LOCvalid)
                  ReportType.config(fg="light slate blue")

               if LOCSponsor == "Account ID":
                  clear()
                  ReportType.config(text=LOCAgency)
                  ReportType.config(fg="orchid1")
                  # Plusbutton.grid(row=15, column=1, columnspan=6, pady=5, padx=5, sticky="nsew")
               print(LOCAgency)
            except:
               IndexError
            ProjectIDregex = r'[50]\d{5}'
            matche = re.search(ProjectIDregex, ShortName)
            if matche:
               ProjectID.config(text=matche.group())
            sheetnames = Df3.sheet_names
            print(sheetnames)
            WRSsearch = re.compile(r'WRS_Report_P1_')
            WRSP3tag = re.compile(r'WRS_Report_P3_')
            WRStag = [name for name in sheetnames if WRSsearch.search(name)]
            WRStag2 = [name for name in sheetnames if WRSP3tag.search(name)]

            if WRStag:
               clear()
               df = pd.read_excel(FileSelection)
               DetailButton.grid()
               ReportType.config(text="P1:")
               ReportType.config(fg="yellow")
               
               
               # max_length = max(df.apply(len))
               # df = df.apply(lambda x: x if len(x) == max_length else x + [None] * (max_length - len(x)))
               # df_nan = df.applymap(lambda x: np.nan if x == '' or pd.isnull(x) else x)
               print(df.tail())
               icon = file_icons.get(file_extension, default_icon)
               DocumentType.config(image=icon)
               SponsorName = df.iloc[12,1]
               AsofDate = df.iloc[0,0]
               Pinvestigator = df.iloc[4,3]
               DocStatus.config(text=SponsorName)
               OUCpanda = df.iloc[4,5]
               Alltime = df.iloc[7,3]
               FArateinto = df.iloc[8,5]
               Awardtotal = df.iloc[47,2]
               TotalIDC = df.iloc[46,2]
               TotalDC = df.iloc[45,2]
               WRSKASH = df.iloc[15,4]
               OtherReceiveables = df.iloc[16,4]
               PTDTotal = df.iloc[47,4]
               PTDIDC = df.iloc[46,4]
               PTDDC = df.iloc[45,4]
               clear()
               # Placeholder.config(text=AsofDate)
               P1Summary.config(text=f"Summary - {AsofDate}")
               P1Summary.config(fg="orange")
               PIname.insert(0, str(Pinvestigator))
               PIname.config(state="readonly")
               POPInfo.config(text=Alltime)
               OUCcode.insert(0,str(OUCpanda))
               OUCcode.config(state="readonly")
               FaInfo.config(text=FArateinto)
               FaInfo.config(fg="yellow")
               Budgetinfo.insert(0,str('{:,.2f}'.format(Awardtotal)))
               Budgetinfo.config(state="readonly")
               WRSCashB.insert(0,str('{:,.2f}'.format(WRSKASH)))
               WRSCashB.config(state="readonly")
               WRSRecivables.insert(0,str('{:,.2f}'.format(OtherReceiveables)))
               WRSRecivables.config(state="readonly")
               PTDXpenses.insert(0,str('{:,.2f}'.format(PTDTotal)))
               PTDXpenses.config(state="readonly")
               WRSTotalExpenses.insert(0,str('{:,.2f}'.format(PTDDC)))
               WRSTotalExpenses.config(state="readonly")
               WRSIndirectcost.insert(0,str('{:,.2f}'.format(PTDIDC)))
               WRSIndirectcost.config(state="readonly")
               # WRSIDbudget.delete(0, tk.END)
               WRSIDbudget.insert(0,str('{:,.2f}'.format(TotalIDC)))
               WRSIDbudget.config(state="readonly")
               WRSDbudget.insert(0,str('{:,.2f}'.format(TotalDC)))
               WRSDbudget.config(state="readonly")
               if WRSKASH < 0:
                  CashB.config(fg="orange")
               if WRSKASH > 0:
                  CashB.config(fg="red")
               if WRSKASH == 0:
                  CashB.config(fg="SeaGreen1")   
               if abs(OtherReceiveables) == abs(Awardtotal):
                  Receivables.config(fg="SeaGreen1")
               if OtherReceiveables == 0:
                  Receivables.config(fg="red")
               if OtherReceiveables < 0:
                  Receivables.config(fg="orange")
               print(abs(WRSKASH))
               print(abs(Awardtotal))
               if PTDTotal > Awardtotal:
                  Totalcost.config(fg="red")
               if PTDTotal < Awardtotal:
                  Totalcost.config(fg="sky blue")
               if PTDDC > TotalDC:
                  DirectExpenses.config(fg="red")
               if PTDDC < TotalDC:
                  DirectExpenses.config(fg="sky blue")
               if PTDIDC > TotalIDC:
                  Indirectcost.config(fg="red")
               if PTDIDC < TotalIDC:
                  Indirectcost.config(fg="sky blue")
               if PTDIDC == TotalIDC:
                  Indirectcost.config(fg="sky blue")


               PTDExpensesFR.insert(0,str('{:,.2f}'.format(PTDTotal)))
               PTDExpensesFR.config(state="readonly")
               print(WRSKASH)
            # if "WRS_Report_P1" in ShortName:
               

            if WRStag2:
               ReportType.config(text="P3: Payroll Transactions")
               ReportType.config(fg="pink")
               Plusbutton.grid(row=15, column=1, columnspan=6, pady=5, padx=5, sticky="nsew")
               ProjectIDregex = r'[50]\d{5}'
               matche = re.search(ProjectIDregex, ShortName)
               if matche:
                  ProjectID.config(text=matche.group())
            # if file_extension in ['xls']:
               # with open(FileSelection, 'r', encoding='utf-8') as file:
                  # xml_content = file.read()
               # df = pd.read_xml(xml_content) #engine='xlrd')
            # print(df)

      progress_var.set(0)
      if not FileSelection:
         # messagebox.showerror("Select A File", "No file selected. Please select a file.")
         FileName1.config(text ="No File Found")
         file_types.clear()
         Dframe.grid_remove()
         DetailButton.grid_remove()
         Inspector.grid_remove()
         DetailBar.grid_remove()
         Plusbutton.grid_remove()

      print(SelectedFile)



def Load2():
   FileSelection2 = filedialog.askopenfilename(filetypes=[("xlsx files", ".xlsx"),("csv files", ".csv"),("xls files", ".xls")])
   global SelectedFile2
   SelectedFile2 = Path(FileSelection2)
   ShortName2 = SelectedFile2.name
   FileName2.config(text=ShortName2)
   print(SelectedFile2)

def get_file_path():
   # global CSV4L
   global file_path
   # Open and return file path
   default_dir = os.path.expanduser("~")
   file_path = filedialog.askdirectory(title="Choose a Destination", initialdir=default_dir)
   # global pathobject
   if not file_path:
      pathobject = defaultlo
      relative = dflon
      os.chdir(pathobject)
   else:
      pathobject = Path(file_path)
      relative = pathobject.name
      os.chdir(pathobject)
      # CSV4L = os.path.join(pathobject, default_name)
      # return CSV4L
   print(relative)
   # print(CSV4L)

   Location.config(text=relative)
   # pdf_scan(file_path)

filecheck = os.path.join(pathobject, default_name)
print(filecheck)


def PDFREADER():
   namecheck = os.path.join(os.getcwd(), OName.get() + ".csv")
   # print(namecheck)
   # if FileName1.cget("text") == "No File Found":
      # messagebox.showerror('Please Choose a File', 'No File Selected!')
   if os.path.exists (namecheck):
      messagebox.showerror('File Already Exist', 'Please Choose a Different FileName')
   
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1")    
         file = SelectedFile
         # file2 = FileSelection

         doc = fitz.open(SelectedFile)
         text_Doc = []
         for pages in range(len(doc)):
            step_size = (pages / len(doc)) * 100
            page = doc[pages]
            text = page.get_text()
            lines = text.split('\n')
            # lines = re.split(r"(\.\d{2}\s*)", text)
            # lines = re.split(r"(?=000\d{6})", text)

            # lines = split(r"(\$\s*-?\d{1,3}(?:,\d{3})*\.\d{2}\s*)", text)

            text_Doc.extend(lines)

            progress_var.set(step_size)
            root.update_idletasks()
            time.sleep(0.1)
         print(text_Doc)
         
         csv_file = OName.get()+".csv"
         
         csv_file_path = os.path.join(os.getcwd(), csv_file)
         # os.path.join(os.getcwd(), csv_file)
         # print(csv_file_path)
         with open(csv_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for line in text_Doc:
               writer.writerow([line])
         progress_var.set(100)
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1") 
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return
      # WireName = OName.get()
      # doc = fitz.open(SelectedFile)
      # for page_num in range(len(doc)):
      #    page = doc[page_num]
      #    text = page.get_text()
   # print(text)


def RETROACTIVITIES():
   namecheck = os.path.join(os.getcwd(), OName.get() + ".csv")
   # print(namecheck)
   # if FileName1.cget("text") == "No File Found":
      # messagebox.showerror('Please Choose a File', 'No File Selected!')
   if os.path.exists (namecheck):
      messagebox.showerror('File Already Exist', 'Please Choose a Different FileName')
   
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1")   
         file = SelectedFile
         
         doc = fitz.open(SelectedFile)
         text_Doc = []
         for pages in range(len(doc)):
            step_size = (pages / len(doc)) * 100
            page = doc[pages]
            text = page.get_text()
            # lines = text.split('\n')
            # lines = re.split(r"(\.\d{2}\s*)", text)
            # lines = re.split(r"(?=\d{9})", text)
            lines = re.split(r"(?=\d{9}(?!.*\f))|(?=NC State University(?!.*\f))", text)
            # lines = re.split(r"(?=\d{9})|(?=NC State University)", text)
            # lines = re.split(r"(?=000\d{6})", text)

            # lines = split(r"(\$\s*-?\d{1,3}(?:,\d{3})*\.\d{2}\s*)", text)

            text_Doc.extend(lines)

            progress_var.set(step_size)
            root.update_idletasks()
            time.sleep(0.1)
         time_pattern = re.compile(r"\b(?:[01]?\d|2[0-3]):[0-5]\d:[0-5]\d\b")
         datepattern = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
         for items in text_Doc:
            try:
               timematch = re.search(time_pattern, items)
               if timematch:
                  Runtime = timematch.group()
                  print(Runtime)
                  break
            except Exception as e:
               print(f"An error occurred: {e}")  
         for items in text_Doc:
            try:
               datematch = re.search(datepattern, items)
               if datematch:
                  Dater = datematch.group()
                  print(Dater)
                  break
            except Exception as e:
               print(f"An error occurred: {e}") 
         
         csv_file = OName.get()+".csv"
         
         csv_file_path = os.path.join(os.getcwd(), csv_file)
         # os.path.join(os.getcwd(), csv_file)
         # print(csv_file_path)
         with open(csv_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for line in text_Doc:
               writer.writerow([line])

         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         sheet = ExcelApp.Workbooks.Open(csv_file_path)
         # sheet.SaveAs(xlsx_file, FileFormat=51) 
         RemoveIndent = '=SUBSTITUTE(A2, CHAR(10), " ")'
         Appender = '=IF(LEFT(B2,3)="Bad",B1&B2,B2)'
         SuperSplitter = '=LEFT(E2, MIN(IF(ISNUMBER(VALUE(MID(E2, ROW($1:$500), 1))), ROW($1:$500))) - 1)'
         NameSplitter = '=TEXTBEFORE(D2,"MTH")'
         DateSplitter = '=TEXTAFTER(D2,"MTH")'
         DateSplitter2 = '=TEXTBEFORE(TEXTAFTER(D2,"MTH")," N")'
         JunkSplitter = '=RIGHT(E2,8)'
         AccountSplitter = '=TEXTSPLIT(J2,"Bad")'
         GroupSplitter = '=TEXTBEFORE(R2,"Good AcctCd")'
         Rebirth = '=TEXTAFTER(D2," N")'
         Ventus = '=TRIM(U2)'

         Retroport = sheet.Sheets(1)

         
         last_row = Retroport.Cells(Retroport.Rows.Count, 1).End(3).Row
         for row in range(last_row, 1, -1):
            if Retroport.Cells(row, 1).Value is None or Retroport.Cells(row, 1).Value == "":
               Retroport.Rows(row).Delete()

         for row in range(last_row, 1, -1):
            Gopack = str(Retroport.Cells(row, 1).Value)
            if "NC State" in Gopack:
               Retroport.Rows(row).Delete()

         

         Retrostart = Retroport.Range("A2")
         RetroEnd = Retroport.Range(Retrostart, Retrostart.End(c.xlDown))
         last_row_number = RetroEnd.Rows(RetroEnd.Rows.Count).Row            
         recordrow = 'B'+str(last_row_number)
         employee = 'C'+str(last_row_number)


         print(last_row_number)
         CleanRetro = Retroport.Range("B2", recordrow)
         EmployCol = Retroport.Range("C2", employee)

         CleanRetro.Formula = RemoveIndent
         CleanRetro.Copy()
         CleanRetro.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         EmployCol.Formula = Appender
         EmployCol.Copy()
         EmployCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         Retroport.Columns(2).Delete()

         for row in range(last_row, 1, -1):
            Gopack2 = str(Retroport.Cells(row, 1).Value)
            if "$" not in Gopack2:
               Retroport.Rows(row).Delete()
         last_row_number = RetroEnd.Rows(RetroEnd.Rows.Count).Row

         delta = 'E'+str(last_row_number)
         iota = 'G'+str(last_row_number)
         sigma = 'F'+str(last_row_number)
         kappa = 'H'+str(last_row_number)
         omega = 'I'+str(last_row_number)
         alpha = 'J'+str(last_row_number)
         lando = 'K'+str(last_row_number)
         phi = 'L'+str(last_row_number)

         DeltaCol = Retroport.Range("E2", delta)
         IotaCol = Retroport.Range("G2", iota)
         SigCol = Retroport.Range("F2", sigma)
         KappaCol = Retroport.Range("H2", kappa)
         OmegaCol = Retroport.Range("I2", omega)
         AlphaCol = Retroport.Range("J2", alpha)
         LandoCol = Retroport.Range("K2", lando)
         PhiCol = Retroport.Range("K2", phi)  
         MuCol = Retroport.Range("L2", phi)  

         CleanRetro = Retroport.Range("B2", recordrow)
         EmployCol = Retroport.Range("C2", employee)
         CleanRetro.TextToColumns(Destination=CleanRetro, DataType=c.xlFixedWidth, FieldInfo=[(0, c.xlTextFormat),(9, 1)])
         # input()
         EmployCol.TextToColumns(Destination=EmployCol, DataType=c.xlFixedWidth, FieldInfo=[(0, c.xlTextFormat),(1,1)])
         DeltaCol.Formula = NameSplitter
         DeltaCol.Copy()
         DeltaCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         SigCol.Formula2 = SuperSplitter
         SigCol.Copy()
         SigCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         IotaCol.Formula = JunkSplitter
         IotaCol.Copy()
         IotaCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         KappaCol.Formula2 = DateSplitter2
         KappaCol.Copy()
         KappaCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         KappaCol.NumberFormat = "m/d/yyyy"
         # KappaCol.Value = KappaCol.Value
         # KappaCol.TextToColumns(Destination=None)
         # input()
         # KappaCol.TextToColumns(Destination=KappaCol, DataType=c.xlFixedWidth, FieldInfo=[(0, c.xlTextFormat),(12, 1)])
         OmegaCol.Value = "N"
         AlphaCol.Formula = Rebirth
         AlphaCol.Copy()
         AlphaCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         # input()
         # OmegaCol.TextToColumns(Destination=OmegaCol, DataType=c.xlFixedWidth, FieldInfo=[(0, c.xlTextFormat),(2,1)])
         # AlphaCol.TextToColumns(Destination=LandoCol, DataType=c.xlDelimited, TextQualifier=c.xlDoubleQuote, ConsecutiveDelimiter=True, Tab=False, Semicolon=False, Comma=False, Space=True, Other=False, OtherChar="")
         # AlphaCol.TextToColumns(Destination=AlphaCol, DataType=c.xlFixedWidth, FieldInfo=[(0, c.xlTextFormat),(4,1)])
         # input()
         LandoCol.Formula2 = AccountSplitter
         PhiCol.Copy()
         PhiCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         Retroport.Columns(12).Insert()
         Retroport.Columns(12).Insert()
         Retroport.Columns(12).Insert()
         LandoCol.TextToColumns(Destination=LandoCol, DataType=c.xlDelimited, TextQualifier=c.xlDoubleQuote, ConsecutiveDelimiter=True, Tab=False, Semicolon=False, Comma=False, Space=True, Other=False, OtherChar="")
         # MuCol.TextToColumns(Destination=MuCol, DataType=c.xlDelimited, TextQualifier=c.xlDoubleQuote, ConsecutiveDelimiter=True, Tab=False, Semicolon=False, Comma=False, Space=False, Other=True, OtherChar=":")
         gamma = 'O'+str(last_row_number)
         GammaCol = Retroport.Range("O2", gamma)
         # Retroport.Columns(15).Insert()
         GammaCol.TextToColumns(Destination=GammaCol, DataType=c.xlDelimited, TextQualifier=c.xlDoubleQuote, ConsecutiveDelimiter=True, Tab=False, Semicolon=False, Comma=False, Space=False, Other=True, OtherChar=":")
         # input()
         Retroport.Columns(17).Insert()
         zeta = 'Q'+str(last_row_number)
         ZetaCol = Retroport.Range("Q2", zeta)
         ZetaCol.Formula2 = GroupSplitter
         ZetaCol.Copy()
         ZetaCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         
         rho = 'R'+str(last_row_number)
         RhoCol = Retroport.Range("R2", rho)
         RhoCol.TextToColumns(Destination=RhoCol, DataType=c.xlFixedWidth, FieldInfo=[(0, c.xlTextFormat),(18, 1)])
         Retroport.Columns(20).Insert()
         Psi = 'S'+str(last_row_number)
         PsiCol = Retroport.Range("S2", Psi)
         PsiCol.TextToColumns(Destination=PsiCol, DataType=c.xlDelimited, TextQualifier=c.xlDoubleQuote, ConsecutiveDelimiter=True, Tab=False, Semicolon=False, Comma=False, Space=True, Other=False, OtherChar="")

         beta = 'U'+str(last_row_number)
         BetaCol = Retroport.Range("U2", beta)
         BetaCol.NumberFormat = "@"
         BetaCol.Replace(What="$", Replacement="", LookAt=2)
         Fin = 'V'+str(last_row_number)
         FinCol = Retroport.Range("V2", Fin)
         FinCol.Formula = Ventus
         FinCol.Copy()
         FinCol.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)


         FinCol.Replace(What=" ", Replacement="", LookAt=2, SearchOrder=1, MatchCase=False, SearchFormat=False, ReplaceFormat=False)
         FinCol.NumberFormat = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
         FinCol.TextToColumns(Destination=None)
         # BetaCol.TextToColumns(Destination=BetaCol, DataType=c.xlDelimited, TextQualifier=c.xlDoubleQuote, ConsecutiveDelimiter=True, Tab=False, Semicolon=False, Comma=False, Space=False, Other=True, OtherChar="$")
         
         # input()
         Retroport.Columns(21).Delete()
         Retroport.Columns(18).Delete()
         Retroport.Columns(11).Delete()
         Retroport.Columns(10).Delete()
         # input()
         Retroport.Columns(5).Delete()
         Retroport.Columns(4).Delete()
         Retroport.Columns(1).Delete()
         # input()       
         
         # LandoCol.TextToColumns()
         # Retroport.Columns(19).Insert()
         # Retroport.Columns(21).Insert()
         
         Retroport.Rows(1).Resize(2).Insert()
         Retroport.Rows(1).Resize(1).Insert()
         Retroport.Cells(2, "A").Value = "Report ID:"
         Retroport.Cells(3, "A").Value = "Emplid"
         Retroport.Cells(3, "B").Value = "Rcd #"
         Retroport.Cells(3, "C").Value = "Employee Name"
         Retroport.Cells(3, "D").Value = "Redist Seq#"
         Retroport.Cells(3, "E").Value = "End Date"
         Retroport.Cells(3, "F").Value = "Off?"
         Retroport.Cells(3, "G").Value = "Page#"
         Retroport.Cells(1, "H").Value = "Run Date:"
         Retroport.Cells(2, "H").Value = "Run Time:"
         Retroport.Cells(3, "H").Value = "Line #"
         Retroport.Cells(3, "I").Value = "Acct_Cd"
         Retroport.Cells(3, "J").Value = "Bad"
         Retroport.Cells(3, "K").Value = "Group"
         Retroport.Cells(3, "L").Value = "GroupCD"
         Retroport.Cells(3, "M").Value = "Good AcctCD"
         Retroport.Cells(3, "N").Value = "Type"
         Retroport.Cells(3, "O").Value = "Amount"
         Retroport.Cells(2, "I").Value = Runtime
         Retroport.Cells(1, "I").Value = Dater
         Retroport.Cells(2, "B").Value = "NCDS0024"
         Retroport.Columns.AutoFit()
         Retroport.Cells(1, "A").Value = "NC State University: INVALID RETROACTIVE DISTRIBUTION ACCT_CD REPORT"
         sheet.Save()

         # amphersan = "@"
         # rplace = ""
         # DeltaCol.Replace(What:="'", Replacement:="", LookAt=2, SearchOrder=1, MatchCase=False, SearchFormat=False, ReplaceFormat=False)
         # for cell in DeltaCol:
         #     if cell.HasFormula:  # Check if the cell contains a formula
         #         cell.Formula = cell.Formula.replace("@", "")
         # WireName = OName.get()
         # doc = fitz.open(SelectedFile)
         # for page_num in range(len(doc)):
         #    page = doc[page_num]
         #    text = page.get_text()
      # print(text)
         progress_var.set(100)
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         progress_var.set(100)
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

def Invictus():
   # namecheck = os.path.join(os.getcwd(), OName.get() + ".csv")
   # print(namecheck)
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   # elif os.path.exists (namecheck):
      # messagebox.showerror('File Already Exist', 'Please Choose a Different FileName')
   
   else:   
      file = SelectedFile
      doc = fitz.open(SelectedFile)
      text_Doc = []
      for pages in range(len(doc)):
         step_size = (pages / len(doc)) * 100
         page = doc[pages]
         text = page.get_text()
         lines = text.split('\n')
         text_Doc.extend(lines)

         progress_var.set(step_size)
         root.update_idletasks()
         time.sleep(0.1)
      print(text_Doc)
      
      csv_file = OName.get()
      
      csv_file_path = os.path.join(os.getcwd(), csv_file)
      # os.path.join(os.getcwd(), csv_file)
      # print(csv_file_path)
      with open(csv_file, 'w', newline='') as csvfile:
         writer = csv.writer(csvfile)
         for line in text_Doc:
            writer.writerow([line])


      ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
      ExcelApp.visible = True
      sheet = ExcelApp.Workbooks.Open(csv_file_path)
      Inviport = sheet.Sheets(1)
      Cert = Inviport.Range('21:25')
      Cert.EntireRow.Delete()
      Cert2 = Inviport.Range('57:58')
      Cert2.EntireRow.Delete()
      Inviport.Rows(2).Delete()
      # used_range = Inviport.UsedRange
      column_values = Inviport.Columns(1).Value

      # RL2 = Frontsheet.Range(ReportLength)
      # print(RLRange)

      for i, cell_value in enumerate(column_values):
         # global found_row
         if cell_value[0] == 'CURRENT':   
            found_row1 = i + 1
            # return found_row
            print(found_row1)
      CurrentRow = 'A'+str(found_row1)   

      for i, cell_value in enumerate(column_values):
         # global found_row
         if cell_value[0] == 'TOTAL':   
            found_row = i + 1
            # return found_row
            # print(found_row2)
      TotalRow = 'A'+str(found_row2) 

      CatRange = Inviport.Range(CurrentRow,TotalRow)

      MoneyPattern = re.compile(r'\b\d+(?:\.\d+)?\b')
      Budgetcat = []
      for row in CatRange:
         for cell in row:
            if cell.Value:
               matches = MoneyPattern.findall(str(cell.Value))
               if matches:
                  Budgetcat.extend(matches)

      print(Budgetcat)
   Status.config(text="DONE!")
   Status.config(foreground="VioletRed1")              
   root.event_generate("<<ProcessCompleted>>", when="tail")

def pdf_scan():
   namecheck = os.path.join(os.getcwd(), OName.get() + ".csv")
   print(namecheck)
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
      return
   elif os.path.exists (namecheck):
      messagebox.showerror('File Already Exist', 'Please Choose a Different FileName')
      return
   
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1") 
         file = SelectedFile               
         WireName = OName.get()
         doc = fitz.open(SelectedFile)
         extracted_text = ""
         amount_list = []
         unique_list = []
         vendor_list = []
         invoice_list = []
         agency_list = []
         progress_var.set(0)
         for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            # lines = text.split('\n')
            # filtered_lines = [line.strip() for line in lines if line.strip()]
            # extracted_text.extend(filtered_lines)
            # extracted_text = "\n".join(extracted_text)
            extracted_text += text.replace('\n', ' ')
            progress_var.set(page_num)
            root.update_idletasks()
            time.sleep(0.1)
         pattern = r"Credit Amount:\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?|\.\d{2})"
         # vendor = r"Company Name:\s*(.+)"
         uniqueID = r"Unique ID:\s{1}(\S+)"
         agency = r"Company Name:\s*(.*?)\s*Standard"
         addy = r"Address 1:\s*(.{11})"
         # addy =  r"(\d+)\s+Originator's Account Number:"
         # addy = r"(.*?)(\S+)\s+(\S+)\s+Originator's Account Number:"
         # r"(?<=\bOriginator's Account Number:)\s*(\S+)\s+(\S+)"
         # vendor2 = r"Unique ID:\s{1}(\S+)"
         matches = re.findall(pattern, extracted_text)
         oneofone = re.findall(uniqueID, extracted_text)
         agencies = re.findall(agency, extracted_text)
         addresses = re.findall(addy, extracted_text)
         unique_list.extend(oneofone)
         # This one works
         # vendor3 = r"(?<=Unique ID:)(?:\s*\S+){1}\s*(\S+)"
         # vendor3 = r"(?<=Unique ID:)(?:\s*\S+){1}\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)"
         # This ONE WORKS as of 4/4
         # vendor3 =r"Unique ID:\s*(?:.*?\s+){1}(.*?)(?=\s{5})"
         # vendor3 = r"Unique ID:\s*(?:.*?\s+){1}(.*?)(?=\s*\d{1,2}/\d{2}/\d{4})"
         vendor3 = r"Unique ID:\s*(?:.*?\s+){1}(.*?)(?:\d{1,2}/\d{1,2}/\d{4})"
         # merger = re.compile(vendor3)
         matchys = re.findall(vendor3, extracted_text)
         print(matchys)
         # for match in matchys:

         vendor_list.extend(matchys)
         agency_list.extend(agencies)
         # for i in range(len(addresses)):
            # if i % 2 != 0:
               # agency_list.extend(addresses) 
         agency_list.extend(addresses[::2])
         
         # Vendies = re.findall(vendor3, extracted_text)
         # vendor_list.extend(Vendies)
         # for match in oneofone:
            # Create a pattern based on the current match
            # pattern_combined = r'\b{}\t(\S+)\b'.format(match)
            # vendor2 = re.search(pattern_combined, extracted_text)
            # vendor_list.extend(vendor2)
          
         
         # whoareu = re.findall(vendor2, extracted_text)

         # for match in whoareu:
            # uniqueID = match.group(1)
            # next_row = match.group(2)
         # if whoareu:
            # unique_id = whoareu.group(1)
            # next_row_text = whoareu.group(2)
         foundinvoices = re.findall(r'\bCNG\d{7}\b', extracted_text)
         WeirdInvoices = r'NCR\d{7}'
         weirdvoices = re.findall(WeirdInvoices, extracted_text)
         
            # print(extracted_text)
         

         amount_list.extend(matches)
         # vendor_list.extend(match_data)
         invoice_list.extend(foundinvoices)
         invoice_list.extend(weirdvoices)

         csv_file = WireName + ".csv"
         
         csv_file_path = os.path.join(os.getcwd(), csv_file)
         # os.path.join(os.getcwd(), csv_file)
         # print(csv_file_path)
         with open(csv_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            # Write header if needed
            # writer.writerow(['Column1', 'Column2', ...])
            writer.writerow(['Wire Amounts', 'Agency', 'Vendor Description', 'Found Invoices', 'Total # of Wires', 'Total Wire Amount'])
            zipped = list(itertools.zip_longest(amount_list, agency_list, vendor_list, invoice_list, fillvalue=None))

            for val1, val2, val3, val4 in zipped:
                writer.writerow([val1, val2, val3, val4])

               # for vendor in vendor_list:
               #    writer.writerow([vendor])
         # print(Wiresheet)
         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         sheet = ExcelApp.Workbooks.Open(csv_file_path)
         Wiresheet = sheet.Sheets(1)
         Wiresheet.Columns(1).NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         Wiresheet.Columns(2).AutoFit()
         # Awrapper = Wiresheet.Range("B:B")
         # Awrapper.AutoFit()
         # Awrapper.WrapText = True
         # Awrapper.ColumnWidth = 40
         VenWrapper = Wiresheet.Range("C:C")
         # VenWrapper.AutoFit()
         VenWrapper.WrapText = True
         VenWrapper.ColumnWidth = 88
         # VenWrapper2 = 
         # VenWrapper.Rows.Autofit()
         Wiresheet.Columns(4).AutoFit()
         Wireone = Wiresheet.Range("A2")
         WireEnd = Wiresheet.Range(Wireone, Wireone.End(c.xlDown))
         # WireEnd.Rows.Autofit()
         Wiresheet.Rows.AutoFit()

         Wiresheet.Cells(2, "E").Value = len(WireEnd)
         Wiresheet.Columns(5).AutoFit()
         Wsum = ExcelApp.WorksheetFunction.Sum(WireEnd)
         Wiresheet.Cells(2, "F").Value = Wsum
         Wiresheet.Cells(2, "F").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         Wiresheet.Columns(6).AutoFit()
         Wiresheet.Activate()
         # ExcelApp.SendKeys("^A")
         # ExcelApp.SendKeys("%HOA")
         print(len(WireEnd))
         print(extracted_text)
         print(vendor_list)
         progress_var.set(100)
         sheet.Save()
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         root.event_generate("<<ProcessCompleted>>", when="tail")
         win32com.client.Dispatch("WScript.Shell").AppActivate("ExcelApp")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return
   #    # progress_var.set(0)
   
   # time.sleep(0.1)
   # sheet.Save()
   # sheet.Close(True)

   # print(email_list)
      # for match in matches:
         # print(match)
         # email_list.append(match)

   
   # print(amount_list)
   # print(len(WireEnd))


Loadb1= tk.Button(Wid_frame, text="Load  ", command=Load, image=openfolder, compound="right")
Loadb1.grid(row=2, column=1, padx=10, pady=10, sticky="se")

def open_workbook():
   if SelectedFile:
      os.startfile(SelectedFile)
   elif FS2:
      os.startfile(FS2)

Inspector = tk.Button(Wid_frame, image=inspect_icon, command=open_workbook)
Inspector.grid(row=3, column=1, padx=10, pady=10, sticky="se")
Inspector.grid_remove()

# Loadb2= ttk.Button(Slot2, text="Load", command=Load)
# Loadb2.grid(row=4, column=3)

# R1 = Radiobutton(FileType, text="Spreadsheet")
# R1.grid(row=0, column=1, padx=20, pady=10, sticky="nse")
# R2 = Radiobutton(FileType, text="PDF")
# R2.grid(row=0, column=2, padx=20, pady=10, sticky="nsw")

start_index = 6
# global Lo
# Lo = StringVar()
# Lo = "DOWNLOADS"
current_directory = Path.cwd()
parent_directory = current_directory.parent
current_directory_name = current_directory.name
current_directory_stem = current_directory.stem
current_directory_suffix = current_directory.suffix

# name = path.name()
# abpath = os.path.relpath(file_path)
# print(relative)
# print (name)


# Reportype = Reports.get()
# print(Reportype)

# Outputname = FileName.get()
#print(c.__dicts__)
global found_row
def Autodraw():
   # if Reports == 
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1") 
         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         Drawsheet = ExcelApp.Workbooks.Open(SelectedFile)
         Frontsheet = Drawsheet.Sheets("Sheet1")
         Frontsheet.Cells.UnMerge()
         Frontsheet.Rows(8).Select()
         ExcelApp.ActiveWindow.FreezePanes = True
         Frontsheet.Columns(4).Insert()
         Frontsheet.Cells(7, "N").Value = "ASAP Balance"
         ASAPCell = Frontsheet.Cells(7, "M")
         # ASAPCell.Interior.Color = 192 
         ASAPCell.Font.Color = 192
         DrawCell = Frontsheet.Cells(7, "O")
         DrawCell.Font.Color = 192
         Frontsheet.Cells(7, "P").Value = "Draw Amount"
         E_range = Frontsheet.Columns("F")
         E_range.NumberFormat = 'mm-dd-yy'
         E_range.TextToColumns(Destination=E_range, DataType=1, TextQualifier=1, Tab=False, Semicolon=False, Comma=False, Space=False, Other=False,)

         BalanceFormula = '=IF(A10="Project Reference Total:",ROUND(G10-J10,2),"")'
         DrawForumula = '=IF(I10<0,N10,IF(AND(K10<0,N10>0,F8>TODAY()-120,D10<>""),K10*-1,""))'
         # (original) DrawForumula = '=IF(AND(K10<0,N10>0,F8>TODAY()-120,I10>0,D10<>""),K10*-1,"")'
         # NSFDrawFormula = 
         ProjectIDFormula = '=IF(A10="Project Reference Total:",A8,"")'
         CashBalanceFormula = '=XLOOKUP(A2,Sheet1!D:D,Sheet1!K:K,"",0)'
         DrawAmountFormula = '=XLOOKUP(A2,Sheet1!D:D,Sheet1!P:P,"",0)'
         DifferenceFormula = '=IF(AND(C2<>"",C2<>0),B2+C2,"")'

         # ReportLength = Frontsheet.Columns(1)
         # RLEnd = Frontsheet.Range(ReportLength,ReportLength.End(c.xlDown))
         # RLRange = Frontsheet.Range(ReportLength,RLEnd)
         # RT = 'Report Total'
         used_range = Frontsheet.UsedRange
         column_values = used_range.Columns(1).Value

         # RL2 = Frontsheet.Range(ReportLength)
         # print(RLRange)

         for i, cell_value in enumerate(column_values):
            # global found_row
            if cell_value[0] == 'Report Total:':   
               found_row = i + 1
               # return found_row
               print(found_row)
         
         FinalASAPRow = 'N'+str(found_row)
         FinalDrawRow = 'P'+str(found_row-3)
         FinalPIDRow = 'D'+str(found_row)
         DSPIDend = 'A'+str(found_row)
         TotalRange = str(found_row-2)
         ReportCB = 'K'+str(found_row)
         RCB = Frontsheet.Range(ReportCB).Value

         # TotalRow = found_row+


         PID = Frontsheet.Range('D10', FinalPIDRow)
         # PIDend = Frontsheet.Range(PID,PID.End(c.xlDown))
         # PIDrange = Frontsheet.Range(PID, PIDend)
         # last_row = PIDend.Row
         # last_column = PIDend.Column
         # FinalCoordinate = last_row, last_column
         PID.Formula = ProjectIDFormula
         IDColumn = Frontsheet.Columns(4).Copy()
         IDRange = Frontsheet.Columns(4)
         IDRange.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         IDRange.TextToColumns(Destination=None)
         PID.NumberFormat = '000000'
         

         print(TotalRange)
         ASAP_range = Frontsheet.Range('N10', FinalASAPRow)
         # Mend_range = Frontsheet.Range(ASAP_range, )
         # Mcolumn_range = Frontsheet.Range(date_range, Mend_range)
         ASAP_range.Formula = BalanceFormula
         # M_range = Frontsheet.Columns("N")
         ASAP_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         
         draw_range = Frontsheet.Range("P10", FinalDrawRow)
         # Oend_range = Frontsheet.Range(draw_range, draw_range.End(c.xlDown))
         # Oend_rangeD = draw_range.End(c.xlDown).End(c.xlToRight)
         # Ocolumn_range = Frontsheet.Range(draw_range, Oend_range)
         draw_range.Formula = DrawForumula
         draw_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         DrawColumn = Frontsheet.Columns(16).Copy()
         PasteRange = Frontsheet.Columns(16)
         PasteRange.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         Frontsheet.Cells(TotalRange,'O').Value = "Total"
         sum_value = ExcelApp.WorksheetFunction.Sum(draw_range)
         DrawRangeCoordinates = "A2:"+DSPIDend
         # Range2sum = Frontsheet.Range(DrawRangeCoordinates)
         # print(DrawRangeCoordinates)
         # DrawTotal = sum(Range2sum).Value
         target_cell = Frontsheet.Range("P"+TotalRange)
         target_cell.Value = sum_value
         target_cell.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
        
         new_sheet = Drawsheet.Sheets.Add(Before = None)
         new_sheet2 = Drawsheet.Sheets.Add(Before = None)

         new_sheet.Name = 'Draw Summary'
         new_sheet2.Name = 'LOC CoverSheet'

         # LOCCS = Drawsheet.Sheets("LOC CoverSheet")
         DS = Drawsheet.Sheets("Draw Summary")

         Projects = Frontsheet.Columns(4).Copy()
         ProjectsDS = DS.Columns(1)
         ProjectsDS.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ProjectsDS.TextToColumns(Destination=None)
         # DS.Range(DrawRangeCoordinates)
         # PDSorted = DS.Range
         DS.Range(DrawRangeCoordinates).Sort(Key1=DS.Range("A2") ,Order1=1, Orientation=1)
         DS.Cells(1, "A").Value = "Project ID"
         DS.Cells(1, "B").Value = "Cash Balance"
         DS.Cells(1, "C").Value = "Draw Amount"
         DS.Cells(1, "D").Value = "Difference"
         DS.Cells(3, "F").Value = "Positive Cash Report"
         DS.Cells(3, "I").Value = "Unavailabe Funding Report"
         DS.Cells(3, "O").Value = "Max Draw Difference"
         DS.Cells(4, "L").Value = "Project ID"
         DS.Cells(3, "L").Value = "Overspent Accounts"
         DS.Cells(4, "M").Value = "Cash Balance"
         # DS.Cells(5, "L").Value = "=I3"
         # DS.Cells(6, "L").Value = "Total Adjustments"
         DS.Cells(4, "I").Value = "Project ID"
         DS.Cells(4, "J").Value = "Cash Balance"
         DS.Cells(4, "O").Value = "Project ID"
         DS.Cells(4, "P").Value = "Amount"



         startrow = DS.Range("A2")
         PIDend = startrow.End(c.xlDown)
         PIDLength = DS.Range(startrow,PIDend)
         LastRow = PIDLength.End(c.xlDown).Row

         print("LAST",LastRow)

         CBFinalRow = 'B'+str(LastRow)
         CB_range = DS.Range('B2', CBFinalRow)
         CB_range.Formula = CashBalanceFormula
         CBColumnget = DS.Columns(2).Copy()
         CBColumn = DS.Columns(2)
         CBColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         CB_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         DrawFR = 'C'+str(LastRow)
         Drange = DS.Range('C2', DrawFR)
         Drange.Formula = DrawAmountFormula
         DRcolumnget = DS.Columns(3).Copy()
         DRColumn = DS.Columns(3)
         DRColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         DRColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         DiffFR = 'D'+str(LastRow)
         DiffRange = DS.Range('D2', DiffFR)
         DiffRange.Formula = DifferenceFormula
         DiffColumnget = DS.Columns(4).Copy()
         DiffColumn = DS.Columns(4)
         DiffColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         DiffColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         PostivePIDFormula = "=FILTER(A:A,B:B>0)"
         PostiveCashBalanceF = "=FILTER(B:B,B:B>0)"
         UnavailableF = '=FILTER(A:A,(C:C="")*(B:B<0))'
         UFCB = '=FILTER(B:B,(C:C="")*(B:B<0))'
         Today = "=TODAY()"
         MaxDrawFormula = '=FILTER(A:A,D:D<0)'
         MaxDrawDiffFormula = '=FILTER(D:D,D:D<0)'
         OverspentActsFormula = '=FILTER(A:A,(C:C=0)*(B:B<0))'
         OverspentCBFormula = '=FILTER(B:B,(C:C=0)*(B:B<0))'
         # UFF = "=IFERROR(INDEX(A:A, SMALL(IF((B:B < 0) * (C:C = ''), ROW(A:A)-MIN(ROW(A:A))+1), ROW(INDIRECT('1:'&COUNTIFS(B:B, '<0', C:C, ''))))), 'No Available Amounts')"

         # workbook = xw.Book(SelectedFile)
         # sheet = workbook.sheets['Draw Summary']

         PIDList = DS.Range("F4").Formula2 = PostivePIDFormula
         PCRStart = DS.Range("F4")

         CoversheetStart = DS.Range("F5")
         CSCBStart = DS.Range("G5")

         PCREnd = PCRStart.End(c.xlDown)
         CSCBEnd = CSCBStart.End(c.xlDown)

         PCRlength = DS.Range(PCRStart,PCREnd)
         
         CSlength = DS.Range(CoversheetStart, PCREnd)

         # CSSBlength = DS.Range(CSCBStart, CSCBEnd)

      #Coversheet Calculations
         # Total Number Of Positive Cash Balance Accounts
         PCRLR = PCRlength.End(c.xlDown).Row
         PCashLastRow = 'F'+str(PCRLR)
         PCashLRange = DS.Range(CSCBStart, PCashLastRow)
         
         PRCtotalRow = 'F'+str(PCRLR+2)
         PRCLabel = DS.Range(PRCtotalRow)
         PCRCol = DS.Columns(6)
         UFT = 'F'+str(PCRLR+3)
         UFTLabel = DS.Range(UFT)
         OT = 'F'+str(PCRLR+4)
         OTLabel = DS.Range(OT)
         TA = 'F'+str(PCRLR+5)
         TALabel = DS.Range(TA)
         

         #Positive Cash Balance Calculations
         PCRCBRow = 'G'+str(PCRLR)
         PCRCBtarget = 'G'+str(PCRLR+2)
         PCRCBtarget2 = DS.Range(PCRCBtarget)
         UFTtarget = 'G'+str(PCRLR+3)
         UFTtarget2 = DS.Range(UFTtarget)
         OSTtarget = 'G'+str(PCRLR+4)
         OSTtarget2 = DS.Range(OSTtarget)
         TAT = 'G'+str(PCRLR+5)
         TAT2 = DS.Range(TAT)

         #CoverSheet Positioning
         PCRTotalNumberAccounts = str(PCRLR-4)
         UFRStartPosition = str(int(PCRTotalNumberAccounts)+11)

         CSUFRAccounts = 'C'+str(UFRStartPosition)
         CSUFRCB = 'D'+str(UFRStartPosition)
         # print(PCRLR)
         print(PCRTotalNumberAccounts)
         print("UFR Start:" + UFRStartPosition)
         

         PRCLabel.Value = "Positive Cash Total"
         PRCLabel.Font.Bold = True
         UFTLabel.Value = "Unavailabe Funding Total"
         UFTLabel.Font.Bold = True
         TALabel.Value = "Total Adjustments"
         TALabel.Font.Bold = True
         OTLabel.Value = "Overspent Total"
         OTLabel.Font.Bold = True

        
         CBlist = DS.Range("G4").Formula2 = PostiveCashBalanceF

      # Unavailabe Funding Report
         # UFR Project IDs
         UF = DS.Range("I5").Formula2 = UnavailableF
         UFPIDStart = DS.Range("I5")
         UFPIDRange = DS.Range(UFPIDStart, UFPIDStart.End(c.xlDown))

         UnavailableFCB = DS.Range("J5").Formula2 = UFCB
         UFRstart = DS.Range("J5")
         UFRend = UFRstart.End(c.xlDown)
         UFRrange = DS.Range(UFRstart, UFRend)
         UFRLR = UFRrange.End(c.xlDown).Row
         # Total Number of Unavialbe Funding Accounts
         UnavailableFTotalAccounts = str(UFRLR-4)
         

         OverspentActs = DS.Range("L5").Formula2 = OverspentActsFormula
         OverspentPIDStart = DS.Range("L5")
         # OverspentPIDEnd = OverspentPIDStart.End(c.xlDown)
         OverspentPIDRange = DS.Range(OverspentPIDStart, OverspentPIDStart.End(c.xlDown))
         
         OverspentCB = DS.Range("M5").Formula2 = OverspentCBFormula
         OverspentCBStart = DS.Range("M5")
         OverspentCBEnd = OverspentCBStart.End(c.xlDown)
         OverspentCBRange = DS.Range(OverspentCBStart, OverspentCBStart.End(c.xlDown))

         OVSA = OverspentPIDRange.End(c.xlDown).Row
         OVSAtotalAccts = str(OVSA-4)
         OVSAStartposition = str(int(UnavailableFTotalAccounts)+int(UFRStartPosition))
         print("OVSA", OVSAStartposition, OVSAtotalAccts)
         CSOVSA = 'C'+str(OVSAStartposition)
         CSOVSCB = 'D'+str(OVSAStartposition)

         MaxDrawActs = DS.Range("O5").Formula2 = MaxDrawFormula
         MaxDrawPIDStart = DS.Range("O5")
         MaxDrawPIDEnd = MaxDrawPIDStart.End(c.xlDown)

         print("MDAR ROW", len(DS.Range(MaxDrawPIDStart, MaxDrawPIDEnd)))
         if len(DS.Range(MaxDrawPIDStart, MaxDrawPIDEnd)) == 1:
            MaxDrawActRange = MaxDrawPIDStart
            MDAR = 1
            MDARtotalAccts = str(MDAR)
         else:
            MaxDrawActRange = DS.Range(MaxDrawPIDStart, MaxDrawPIDStart.End(c.xlDown))
            MDAR = MaxDrawActRange.End(c.xlDown).Row
            MDARtotalAccts = str(MDAR-4)

         MaxDrawDiff = DS.Range("P5").Formula2 = MaxDrawDiffFormula
         MaxDrawDiffStart = DS.Range("P5")
         MaxDrawDiffEnd = MaxDrawDiffStart.End(c.xlDown)
         if MaxDrawDiffEnd.Row > LastRow:
            MaxDrawDiffRange = MaxDrawDiffStart
         else:
            MaxDrawDiffRange = DS.Range(MaxDrawDiffStart, MaxDrawDiffStart.End(c.xlDown))
         print(MaxDrawPIDEnd.Row)
         # input()
         

         # MDAR = MaxDrawActRange.End(c.xlDown).Row
         
         MDARStartpostion = str(int(UnavailableFTotalAccounts)+int(OVSAtotalAccts)+int(UFRStartPosition))
         CSMDAR = 'C'+str(MDARStartpostion)
         CSMDCBR = 'D'+str(MDARStartpostion)
         print("MDAR Total:" + MDARtotalAccts)

         # UFR Cash Balances
         
         # print(UFRLR)
         print(UnavailableFTotalAccounts)
         LOCTotalEnding = str(int(PCRTotalNumberAccounts)+int(UnavailableFTotalAccounts)+int(OVSAtotalAccts)+int(MDARtotalAccts) + 10)
         print("LOCEnding:",LOCTotalEnding)
         UFCBColumn = DS.Columns(10)
         UFCBColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'  
         OverCBColumn = DS.Columns(13)
         OverCBColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         MaxDrawColumn = DS.Columns(16)
         MaxDrawColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         PCRCol.AutoFit()
         PCRrange = DS.Range("G4",PCRCBRow)
         PCTotal = ExcelApp.WorksheetFunction.Sum(PCRrange)
         UFTTotal = ExcelApp.WorksheetFunction.Sum(UFRrange)

         OVSrange = DS.Range(OverspentCBStart, OverspentCBEnd)
         OVSPTotal = ExcelApp.WorksheetFunction.Sum(OVSrange)

         MaxDRange = DS.Range(MaxDrawDiffStart, MaxDrawDiffEnd)
         MaxDTotal = ExcelApp.WorksheetFunction.Sum(MaxDRange)

         OSPNTotal = float(OVSPTotal + MaxDTotal)

         PCRCBtarget2.Value = PCTotal
         UFTtarget2.Value = UFTTotal
         OSTtarget2.Value = OSPNTotal
         TotalAdjRange = DS.Range(PCRCBtarget2,OSTtarget2)

         # Total Adjusmtments 
         TotalAdj = ExcelApp.WorksheetFunction.Sum(TotalAdjRange)
         CSTAdj = float(TotalAdj * -1)

         print(CSTAdj)
         TAT2.Value = TotalAdj
         # sheet.range('I5').formula2 = "=FILTER(A:A,(B:B<0)*(C:C=''))"
         DSCBColumn = DS.Columns(7)
         DSCBColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         # PList = DS.Range("F4").CurrentRegion
         # spilled_values = PList.Value
         # return spilled_values
         # PListFull = DS.Range(PList, PList.End(c.xlDown))
         # PList.Formula = PostivePIDFormula
         new_sheet2.Name = 'LOC CoverSheet'
         LOC = Drawsheet.Sheets("LOC CoverSheet")
         LOC.Cells(1, "A").Value = "LOC CoverSheet"
         LOC.Cells(2, "A").Formula = Today
         LOC.Cells(9, "A").Value = "Adjustments"
         LOC.Cells(3, "C").Value = "Sponsor:"
         LOC.Cells(4, "C").Value = "Period:"
         LOC.Cells(5, "C").Value = "FM (Desk):"
         LOCcolumn = LOC.Columns(1)
         LOCcolumn.AutoFit()
         LOC.Cells(7, "A").Value = "LOC Report Cash Balance:"
         LOC.Cells(9, "E").Value = "Comments"
         
         LOC.Cells(4, "D").Value = '=Sheet1!A4'
         CommentWidth = LOC.Columns(4)
         CommentWidth.AutoFit()
         LOC.Cells(5, "D").Value = '=Sheet1!A6'
         
         LOC.Cells(3, "D").Value = '=Sheet1!A8'
         LOC.Cells(7, "D").Value = RCB
         LOC.Cells(7, "D").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         LOC.Cells(9, "D").Value = CSTAdj
         LOC.Cells(9, "D").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
      
         # Positive Cash Balance -> Coversheet Transfer
         PCRAccounts = CSlength.Copy()
         CSCBalances = PCashLRange.Copy()
         CoversheetAccounts = LOC.Range("C11")
         CoversheetAccounts.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False

         UFRAccounts = UFPIDRange.Copy()
         UFRCoversheet = LOC.Range(CSUFRAccounts)
         UFRCoversheet.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False

         UFRCashbalance = UFRrange.Copy()
         UFRCBCS = LOC.Range(CSUFRCB)
         UFRCBCS.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False

         OVSAccounts = OverspentPIDRange.Copy()
         OVSCoversheet = LOC.Range(CSOVSA)
         OVSCoversheet.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False

         OVSCashBalance = OverspentCBRange.Copy()
         OVCBCoversheet = LOC.Range(CSOVSCB)
         OVCBCoversheet.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False

         MaxAccounts = MaxDrawActRange.Copy()
         MaxACoversheet = LOC.Range(CSMDAR)
         MaxACoversheet.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False

         MaxCB = MaxDrawDiffRange.Copy()
         MaxDiff = LOC.Range(CSMDCBR)
         MaxDiff.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         ExcelApp.CutCopyMode = False
         
         RangeAddress = 'D11:D'+str(LOCTotalEnding)
         TotalCashNeeded = str(int(LOCTotalEnding) + 2)
         print("TCN", TotalCashNeeded)
         TotalFundsBeingRequested = str(int(LOCTotalEnding) + 3)
         Difference = str(int(LOCTotalEnding) + 4)
         SeniorFMApproval = str(int(LOCTotalEnding) + 6)

         TCN = LOC.Cells(TotalCashNeeded, "A")
         TFBR = LOC.Cells(TotalFundsBeingRequested, "A")
         CSD = LOC.Cells(Difference, "A")

         TCN.Value = "Total Cash Needed"
         TFBR.Value = "Total Funds Being Requested"
         CSD.Value = "Difference"
         AdjustmentsTotal = '=D7+D9'
         # LOCReconciliation = '='
         AdjusmentTotalCoord = LOC.Cells(TotalCashNeeded, "D")
         TFBRCoordinates = LOC.Cells(TotalFundsBeingRequested, "D")
         DifferenceCoord = LOC.Cells(Difference, "D")

         AdjusmentTotalCoord.Formula = AdjustmentsTotal
         # Adj2 = float(AdjusmentTotalCoord)
         
         # print(Adj2)

         
         TFBRCoordinates.Value = target_cell
         # TBR2 = float(TFBRCoordinates)
         DifferenceRange = LOC.Range(AdjusmentTotalCoord,TFBRCoordinates)

         LOCReconcile = ExcelApp.WorksheetFunction.Sum(DifferenceRange)


         DifferenceCoord.Value = LOCReconcile
         DifferenceCoord.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         TFBRCoordinates.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

         # DifferenceAmount = str(TFBRCoordinates) + str(AdjusmentTotalCoord)
         # print(DifferenceAmount)

         CSTotal = LOC.Range(RangeAddress)
         for cell in CSTotal:
            cell.Value = cell.Value * -1
         CSTotal.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         progress_var.set(100)
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         root.event_generate("<<ProcessCompleted>>", when="tail")
         win32com.client.Dispatch("WScript.Shell").AppActivate("ExcelApp")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

def Autodraw3():

   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1") 
         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         Drawsheet = ExcelApp.Workbooks.Open(SelectedFile)
         Frontsheet = Drawsheet.Sheets("Sheet1")
         Frontsheet.Cells.UnMerge()
         Frontsheet.Rows(8).Select()
         ExcelApp.ActiveWindow.FreezePanes = True
         Frontsheet.Columns(4).Insert()
         Frontsheet.Cells(7, "N").Value = "PMS Balance"
         new_sheet = Drawsheet.Sheets.Add(After = Drawsheet.Sheets("Sheet1"))
         new_sheet.Name = "Superfund"
         Superfund = Drawsheet.Sheets("Superfund")
         Frontsheet.Rows("1:7").Copy()
         SuperfundHeader= Superfund.Range("A1")
         SuperfundHeader.PasteSpecial()

         new_sheet2 = Drawsheet.Sheets.Add(After = Drawsheet.Sheets("Superfund"))
         new_sheet2.Name = "Other - PMS"
         NonPHS = Drawsheet.Sheets("Other - PMS")
         Frontsheet.Rows("1:7").Copy()
         NonPHS.Range("A1").PasteSpecial()
         

         SuperIDs = Frontsheet.UsedRange
         SuperValues = SuperIDs.Columns(1).Value

            # RL2 = Frontsheet.Range(ReportLength)
         # print(RLRange)

         for i, cell_value in enumerate(SuperValues):
            step_size = (i / len(SuperValues)) * 50
            progress_var.set(step_size)
            root.update_idletasks()
            # global found_row
            if cell_value[0] == 'Report Total:':   
               ReportRow = i + 1
               # return found_row
               print(ReportRow)
               
            
         search_value = "31009"
         ThatOneAccount = "HQ00052110082"
         LOCValue = "LOC#"
         found_rows = []
         found_rows2 = []
         LOCSegments = []
         bins = []
         LOCTotalRow = []
         OriginalRevision = []
         ThatOneRow = []
         lo = 1

         last_row = Frontsheet.Cells(Frontsheet.Rows.Count, 2).End(-4162).Row  # xlUp = -4162

         for i in range(1, last_row + 1):
            LtrsCred = str(Frontsheet.Cells(i, 1).Value) 
            if LOCValue in LtrsCred:
               LOCSegments.append((i, str(LtrsCred)))
            
           

         print(ThatOneRow)
         cutpoints = sorted(LOCSegments)
         
         for cp, label in cutpoints:
            bins.append((lo, cp))
            lo = cp + 1
         bins.append((lo, math.inf)) 

         for i in range(1, last_row + 1):
            cell_value = str(Frontsheet.Cells(i, 2).Value)  # Column B = 2
            if search_value in cell_value:
               found_rows.append(i)
            if ThatOneAccount in cell_value:
               ThatOneRow.append(i)
         print("TOR:", ThatOneRow)

         ranges = []
         if found_rows:
            start = prev = found_rows[0]
            for r in found_rows[1:]:
               if r == prev + 1:
                  prev = r
               else:
                  ranges.append((start, prev)) 
                  start = prev = r
            ranges.append((start, prev))

         adjusted_ranges = [(start - 1, end + 2) for start, end in ranges]


         ranges2 = []
         if ThatOneRow:
            start2 = prev2 = ThatOneRow[0]
            for r in ThatOneRow[1:]:
               if r == prev2 + 1:
                  prev2 = r
               else:
                  ranges2.append((start2, prev2)) 
                  start2 = prev2 = r
            ranges2.append((start2, prev2))

         adjusted_ranges2 = [(start - 1, end + 2) for start, end in ranges2]
         adjusted_ranges2.sort(reverse=True)

         for start, end in adjusted_ranges2:
            Frontsheet.Rows(f"{start}:{end}").Cut(NonPHS.Rows(8))
            Frontsheet.Rows(f"{start}:{end}").Delete()

         print(ranges2)
         print(adjusted_ranges2)
         last_used = Superfund.Cells(Superfund.Rows.Count, 7).End(-4162).Row  # xlUp = -4162

         dest_row = 1 if Superfund.Cells(1, 1).Value is None and last_used == 1 else last_used + 1
         adjusted_ranges.sort(reverse=True)
         first_section = True
         current_bin = None
         for start, end in adjusted_ranges:
            b = bisect.bisect_left([r for r, _ in cutpoints], start)
            if b != current_bin:

               Superfund.Rows(dest_row).Insert()

               idx = bisect.bisect_right([r for r, _ in LOCSegments], start) - 1
               if not first_section and b is not None:
                  Superfund.Rows(dest_row).Insert()
                  Superfund.Cells(dest_row, 1).Value = f"LOC Total:"
                  Superfund.Cells(dest_row, 1).Font.Bold = True
                  # dest.Cells(dest_row, 1).Interior.ColorIndex = 2  # white fill
                  dest_row += 1
               else:
                  first_section = False

               if idx >= 0:
                  label = LOCSegments[idx][1]
               else:
                  label = "Unlabeled Section"


               Superfund.Cells(dest_row, 1).Value = f"{label}"
               # Superfund.Cells(dest_row, 1).WrapText = False
               # Superfund.Cells(dest_row, 1).HorizontalAlignment = -4131
               # Superfund.Cells(dest_row, 1).Interior.ColorIndex =  # light yellow fill
               dest_row += 1
               current_bin = b

            rows_in_block = end - start + 1
            Frontsheet.Rows(f"{start}:{end}").Cut(Superfund.Rows(dest_row))
            Frontsheet.Rows(f"{start}:{end}").Delete()
            dest_row += rows_in_block

         if b is not None:
            Superfund.Rows(dest_row).Insert()
            Superfund.Cells(dest_row, 1).Value = f"LOC Total:"
            Superfund.Cells(dest_row, 1).Font.Bold = True
            # dest.Cells(dest_row, 1).Interior.ColorIndex = 2  # white fill
            dest_row += 1
         Superfund.Columns(1).WrapText = False
         Superfund.Columns(1).HorizontalAlignment = -4131
         Superfund.Columns(2).ColumnWidth = 16
         Superfund.Columns(2).AutoFit()

         last_row2 = Superfund.Cells(Superfund.Rows.Count, 1).End(-4162).Row  # xlUp = -4162

         for i in range(1, last_row2 + 1):
            Lrow = str(Superfund.Cells(i, 1).Value) 
            if "LOC Total:" in Lrow:
               LOCTotalRow.append(i)

         print(LOCTotalRow)

         basecell = 8

         for i in LOCTotalRow:
            SumifsFormula = f"""=SUMIF($A${basecell}:$A${i}, "Project Reference Total:", G{basecell}:G{i})"""
            Ltot = Superfund.Range(f"G{i}").Formula = SumifsFormula
            Superfund.Cells(f"{i}","G").AutoFill(Superfund.Range(f"G{i}:L{i}"))
            # Superfund.Range(f"G{i}:L{i}").Copy()
            # Superfund.Range(f"G{i}:L{i}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
            basecell = i + 1
         
         FiscalMTotal = basecell
         ReportTotal = basecell + 1
         SumifReport = f"""=SUMIF($A$8:$A${FiscalMTotal}, "LOC Total:", G8:G{FiscalMTotal})"""
         Superfund.Range(f"A{FiscalMTotal}").Value = "Fiscal Manager"
         Superfund.Range(f"A{FiscalMTotal}").Font.Bold = True
         Superfund.Range(f"G{FiscalMTotal}").Formula = SumifReport
         Superfund.Cells(f"{FiscalMTotal}","G").AutoFill(Superfund.Range(f"G{FiscalMTotal}:L{FiscalMTotal}"))
         Superfund.Range(f"G{FiscalMTotal}:L{FiscalMTotal}").Copy()
         Superfund.Range(f"G{FiscalMTotal}:L{FiscalMTotal}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         
         Superfund.Range(f"A{ReportTotal}").Value = "Report Total:"
         Superfund.Range(f"G{ReportTotal}:L{ReportTotal}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)

         Superfund.Rows(last_row2 - 1).Copy()
         Superfund.Range(f"A{last_row2}:L{last_row2}").PasteSpecial(Paste=-4122, Operation=-4142, SkipBlanks=False, Transpose=False)
         Superfund.Range(f"A{ReportTotal}:L{ReportTotal}").PasteSpecial(Paste=-4122, Operation=-4142, SkipBlanks=False, Transpose=False)
         Superfund.Range(f"A{FiscalMTotal}:L{FiscalMTotal}").PasteSpecial(Paste=-4122, Operation=-4142, SkipBlanks=False, Transpose=False)

         Superfund.Range(f"A{ReportTotal}").Font.Bold = True
         Superfund.Range(f"G{ReportTotal}:L{ReportTotal}").Font.Bold = True
         Superfund.Range(f"G{ReportTotal}:L{ReportTotal}").NumberFormat = '#,##0.00'
         Superfund.Range(f"G{FiscalMTotal}:L{FiscalMTotal}").Font.Bold = True
         Superfund.Range(f"G{FiscalMTotal}:L{FiscalMTotal}").NumberFormat = '#,##0.00'
         print("Last Row", last_row2)
         

         basecell2 = 8
         last_row3 = Frontsheet.Cells(Frontsheet.Rows.Count, 1).End(-4162).Row
         
         for i in range(1, last_row3 + 1):
            LtrsCred2 = str(Frontsheet.Cells(i, 1).Value) 
            if "LOC Total:" in LtrsCred2:
               OriginalRevision.append(i)

         print("LOC Total:", OriginalRevision)
        
         for i in OriginalRevision:
            SumifsFormula = f"""=SUMIF($A${basecell2}:$A${i}, "Project Reference Total:", G{basecell2}:G{i})"""
            Ltot = Frontsheet.Range(f"G{i}").Formula = SumifsFormula
            Frontsheet.Cells(f"{i}","G").AutoFill(Frontsheet.Range(f"G{i}:L{i}"))
            basecell2 = i + 1
            # Frontsheet.Range(f"G{i}:L{i}").Copy()
         
         FiscalMTotal = basecell2
         ReportTotal = basecell2 + 1
         SumifReport = f"""=SUMIF($A$8:$A${FiscalMTotal}, "LOC Total:", G8:G{FiscalMTotal})"""
         Frontsheet.Range(f"G{FiscalMTotal}").Formula = SumifReport
         Frontsheet.Cells(f"{FiscalMTotal}","G").AutoFill(Frontsheet.Range(f"G{FiscalMTotal}:L{FiscalMTotal}"))
         # input()
         Frontsheet.Range(f"G{FiscalMTotal}:L{FiscalMTotal}").Copy()
         Frontsheet.Range(f"G{FiscalMTotal}:L{FiscalMTotal}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         
         Frontsheet.Range(f"G{ReportTotal}:L{ReportTotal}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         Frontsheet.Columns(4).Delete()
         Frontsheet.Columns.AutoFit()
         Superfund.Columns.AutoFit()
         NonPHS.Columns.AutoFit()

         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         progress_var.set(100)
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

def Autodraw2():
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1") 
         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         Drawsheet = ExcelApp.Workbooks.Open(SelectedFile)
         Frontsheet = Drawsheet.Sheets("Sheet1")
         Frontsheet.Cells.UnMerge()
         Frontsheet.Rows(8).Select()
         ExcelApp.ActiveWindow.FreezePanes = True
         Frontsheet.Columns(4).Insert()
         Frontsheet.Cells(7, "N").Value = "ASAP Balance"
         ASAPCell = Frontsheet.Cells(7, "M")
         # ASAPCell.Interior.Color = 192 
         ASAPCell.Font.Color = 192
         DrawCell = Frontsheet.Cells(7, "O")
         DrawCell.Font.Color = 192
         Frontsheet.Cells(7, "P").Value = "Draw Amount"
         # Frontsheet.Cells(7, "R").Value = "Overspent Draws"
         E_range = Frontsheet.Columns("F")
         E_range.NumberFormat = 'mm-dd-yy'
         E_range.TextToColumns(Destination=None)


         BalanceFormula = '=IF(A10="Project Reference Total:",ROUND(G10-J10,2),"")'
         DrawForumula = '=IF(A10="Project Reference Total:",ROUND(K10*-1,2),"")'
         # '=IF(AND(N10>=0,F8>TODAY()-120,I10>=0,G10>0,D10<>""),ROUND(K10*-1,2),"")'
         # NSFDrawFormula = 
         ProjectIDFormula = '=IF(A10="Project Reference Total:",A8,"")'
         CashBalanceFormula = '=XLOOKUP(A2,Sheet1!D:D,Sheet1!K:K,"",0)'
         DrawAmountFormula = '=XLOOKUP(A2,Sheet1!D:D,Sheet1!P:P,"",0)'
         OverspentFormula = '=IF(AND(N10>0,I10<0,K10<0),N10,"")'
         OSSummaryFormula = '=XLOOKUP(A2,Sheet1!D:D,Sheet1!R:R,"",0)'

         
         used_range = Frontsheet.UsedRange
         column_values = used_range.Columns(1).Value

            # RL2 = Frontsheet.Range(ReportLength)
            # print(RLRange)

         for i, cell_value in enumerate(column_values):
            # global found_row
            if cell_value[0] == 'Report Total:':   
               found_row = i + 1
               # return found_row
               print(found_row)
         
         FinalASAPRow = 'N'+str(found_row)
         FinalDrawRow = 'P'+str(found_row-3)
         FinalPIDRow = 'D'+str(found_row)
         DSPIDend = 'A'+str(found_row)
         TotalRange = str(found_row-2)
         OSTRange = str(found_row-1)
         OSTRange2 = 'P'+str(found_row)
         GTFR = Frontsheet.Range(OSTRange2)
         OSTRange3 = 'P'+str(found_row-1)
         GrandTRange = str(found_row)
         ReportCB = 'K'+str(found_row)
         OverspentRow = 'R'+str(found_row-3)
         RCB = Frontsheet.Range(ReportCB).Value

         # TotalRow = found_row+


         PID = Frontsheet.Range('D10', FinalPIDRow)
         # PIDend = Frontsheet.Range(PID,PID.End(c.xlDown))
         # PIDrange = Frontsheet.Range(PID, PIDend)
         # last_row = PIDend.Row
         # last_column = PIDend.Column
         # FinalCoordinate = last_row, last_column
         PID.Formula = ProjectIDFormula
         IDColumn = Frontsheet.Columns(4).Copy()
         IDRange = Frontsheet.Columns(4)
         IDRange.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
         IDRange.TextToColumns(Destination=None)
         PID.NumberFormat = '000000'

         # Down = win32com.client.constants.xlDown
         

         print(TotalRange)
         ASAP_range = Frontsheet.Range('N10', FinalASAPRow)
         # Mend_range = Frontsheet.Range(ASAP_range, )
         # Mcolumn_range = Frontsheet.Range(date_range, Mend_range)
         ASAP_range.Formula = BalanceFormula
         # M_range = Frontsheet.Columns("N")
         ASAP_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         
         draw_range = Frontsheet.Range("P10", FinalDrawRow)
         Os_range = Frontsheet.Range("R10", OverspentRow)
         GTRange = Frontsheet.Range(FinalDrawRow,OSTRange3)
         # Oend_range = Frontsheet.Range(draw_range, draw_range.End(c.xlDown))
         # Oend_rangeD = draw_range.End(c.xlDown).End(c.xlToRight)
         # Ocolumn_range = Frontsheet.Range(draw_range, Oend_range)
         draw_range.Formula = DrawForumula
         draw_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         DrawColumn = Frontsheet.Columns(16).Copy()
         PasteRange = Frontsheet.Columns(16)
         PasteRange.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)

         MonthlyFormula = f'=K{found_row}*-1'
         MonthlyRange = Frontsheet.Range(f'N{found_row}')
         MonthlyRange.Formula = MonthlyFormula
         GrandFormula = f'=SUM(N10:N{found_row})'
         GrandRange = Frontsheet.Range(f'N{found_row + 1}')
         GrandRange.Formula = GrandFormula
         GrandRange.Copy()
         Dranged = Frontsheet.Range(f'P{found_row}')
         Dranged.PasteSpecial()
         Frontsheet.Range("A1").Select()
         Frontsheet.Columns.AutoFit()
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         progress_var.set(100)
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

   # input()
#    Frontsheet.Cells(TotalRange,'O').Value = "Standard Total"
#    Frontsheet.Cells(OSTRange,'O').Value = "Overspent Total"
#    Frontsheet.Cells(GrandTRange,'O').Value = "Drawsheet Total"
#    Totalcolumn = Frontsheet.Columns(15)
#    Totalcolumn.AutoFit()

#    sum_value = ExcelApp.WorksheetFunction.Sum(draw_range)
#    # OS_value = ExcelApp.WorksheetFunction.Sum(Os_range)=K2861*-1

#    DrawRangeCoordinates = "A2:"+DSPIDend
#    # Range2sum = Frontsheet.Range(DrawRangeCoordinates)
#    # print(DrawRangeCoordinates)
#    # DrawTotal = sum(Range2sum).Value
#    target_cell = Frontsheet.Range("P"+TotalRange)
#    target_cell.Value = sum_value
#    target_cell.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

   
#    OverColumn = Frontsheet.Columns(18).Copy()
#    PasteRange2 = Frontsheet.Columns(18)
#    PasteRange2.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)

#    overspent_range = Frontsheet.Range("R10", OverspentRow)
#    overspent_range.Formula = OverspentFormula
#    overspent_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

#    OS_value = ExcelApp.WorksheetFunction.Sum(Os_range)

#    Otarget_cell = Frontsheet.Range("P"+OSTRange)
#    Otarget_cell.Value = OS_value
#    Otarget_cell.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

#    GrandTotal = ExcelApp.WorksheetFunction.Sum(GTRange)
#    GTFR.Value = GrandTotal
#    GTFR.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
  
#    new_sheet = Drawsheet.Sheets.Add(Before = None)
#    new_sheet2 = Drawsheet.Sheets.Add(Before = None)

#    new_sheet.Name = 'Draw Summary'
#    new_sheet2.Name = 'LOC CoverSheet'
  
#    # LOCCS = Drawsheet.Sheets("LOC CoverSheet")
#    DS = Drawsheet.Sheets("Draw Summary")

#    Projects = Frontsheet.Columns(4).Copy()
#    ProjectsDS = DS.Columns(1)
#    ProjectsDS.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    ProjectsDS.TextToColumns(Destination=None)
#    # DS.Range(DrawRangeCoordinates)
#    # PDSorted = DS.Range
#    DS.Range(DrawRangeCoordinates).Sort(Key1=DS.Range("A2") ,Order1=1, Orientation=1)
#    DS.Cells(1, "A").Value = "Project ID"
#    DS.Cells(1, "B").Value = "Cash Balance"
#    DS.Cells(1, "C").Value = "Draw Amount"
#    DS.Cells(1, "D").Value = "Overspent"
#    DS.Cells(3, "F").Value = "Positive Cash Report"
#    DS.Cells(4, "F").Value = "Project ID"
#    DS.Cells(4, "G").Value = "Cash Balance"
#    DS.Cells(3, "I").Value = "Unavailabe Funding Report"
#    DS.Cells(3, "L").Value = "Overspent Accounts Report"
#    # DS.Cells(4, "L").Value = "Project ID"
#    # DS.Cells(4, "M").Value = "Cash Balance"
#    DS.Cells(4, "I").Value = "Project ID"
#    DS.Cells(4, "J").Value = "Cash Balance"


#    startrow = DS.Range("A2")
#    PIDend = startrow.End(c.xlDown)
#    PIDLength = DS.Range(startrow,PIDend)
#    LastRow = PIDLength.End(c.xlDown).Row

#    # print(LastRow)

#    CBFinalRow = 'B'+str(LastRow)
#    CB_range = DS.Range('B2', CBFinalRow)
#    CB_range.Formula = CashBalanceFormula
#    CBColumnget = DS.Columns(2).Copy()
#    CBColumn = DS.Columns(2)
#    CBColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    CB_range.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

#    DrawFR = 'C'+str(LastRow)
#    Drange = DS.Range('C2', DrawFR)
#    Drange.Formula = DrawAmountFormula
#    DRcolumnget = DS.Columns(3).Copy()
#    DRColumn = DS.Columns(3)
#    DRColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    DRColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

#    OSFinalRow = 'D'+str(LastRow)
#    OSSrange = DS.Range('D2',OSFinalRow)
#    OSSrange.Formula = OSSummaryFormula
#    OSculumnget = DS.Columns(4).Copy()
#    OScolumn = DS.Columns(4)
#    OScolumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    OScolumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'


#    PostivePIDFormula = '=FILTER(A:A,(B:B>0)*(C:C=""))'
#    PostiveCashBalanceF = '=FILTER(B:B,(B:B>0)*(C:C=""))'
#    UnavailableF = '=FILTER(A:A,(C:C="")*(B:B<0))'
#    UFCB = '=FILTER(B:B,(C:C="")*(B:B<0))'
#    OSFilter = '=FILTER(A:A,D:D<>"")'
#    OSCBFilter = '=FILTER(D:D,D:D<>"")'
#    Today = "=TODAY()"
#    # UFF = "=IFERROR(INDEX(A:A, SMALL(IF((B:B < 0) * (C:C = ''), ROW(A:A)-MIN(ROW(A:A))+1), ROW(INDIRECT('1:'&COUNTIFS(B:B, '<0', C:C, ''))))), 'No Available Amounts')"

#    # workbook = xw.Book(SelectedFile)
#    # sheet = workbook.sheets['Draw Summary']

#    PIDList = DS.Range("F5").Formula2 = PostivePIDFormula
#    PCRStart = DS.Range("F5")

#    CoversheetStart = DS.Range("F5")
#    CSCBStart = DS.Range("G5")

#    PCREnd = PCRStart.End(c.xlDown)
#    CSCBEnd = CSCBStart.End(c.xlDown)


#    PCRlength = DS.Range(PCRStart,PCREnd)
   
#    CSlength = DS.Range(CoversheetStart, PCREnd)

#    # CSSBlength = DS.Range(CSCBStart, CSCBEnd)

# #Coversheet Calculations
#    # Total Number Of Positive Cash Balance Accounts
#    PCRLR = PCRlength.End(c.xlDown).Row
#    PCashLastRow = 'F'+str(PCRLR)
#    PCashLRange = DS.Range(CSCBStart, PCashLastRow)
#    PRCtotalRow = 'F'+str(PCRLR+2)
#    PRCLabel = DS.Range(PRCtotalRow)

#    PCRCol = DS.Columns(6)
#    UFT = 'F'+str(PCRLR+3)
#    UFTLabel = DS.Range(UFT)
#    TA = 'F'+str(PCRLR+5)
#    TALabel = DS.Range(TA)
#    OSL = 'F'+str(PCRLR+4)
#    OverspentLabel = DS.Range(OSL)

#    #Positive Cash Balance Calculations
#    PCRCBRow = 'G'+str(PCRLR)
#    PCRCBtarget = 'G'+str(PCRLR+2)
#    PCRCBtarget2 = DS.Range(PCRCBtarget)
#    UFTtarget = 'G'+str(PCRLR+3)
#    UFTtarget2 = DS.Range(UFTtarget)
#    OverspentRow = 'G'+str(PCRLR+4)
#    OverspentTarget = DS.Range(OverspentRow)
#    TAT = 'G'+str(PCRLR+5)
#    TAT2 = DS.Range(TAT)
#    PCRTotalNumberAccounts = str(PCRLR-4)

#    UFRStartPosition = str(int(PCRTotalNumberAccounts)+11)

#    CSUFRAccounts = 'C'+str(UFRStartPosition)
#    CSUFRCB = 'D'+str(UFRStartPosition)
#    # print(PCRLR)
#    print(PCRTotalNumberAccounts)
#    print(UFRStartPosition)
   
#    PRCLabel.Value = "Positive Cash Total"
#    PRCLabel.Font.Bold = True
#    UFTLabel.Value = "Unavailabe Funding Total"
#    UFTLabel.Font.Bold = True
#    OverspentLabel.Value = "Overspent Total"
#    OverspentLabel.Font.Bold = True
#    TALabel.Value = "Total Adjustments"
#    TALabel.Font.Bold = True
   
#    CBlist = DS.Range("G5").Formula2 = PostiveCashBalanceF

# # Unavailabe Funding Report
#    # UFR Project IDs
#    UF = DS.Range("I5").Formula2 = UnavailableF
#    UFPIDStart = DS.Range("I5")
#    UFPIDRange = DS.Range(UFPIDStart, UFPIDStart.End(c.xlDown))

#    UnavailableFCB = DS.Range("J5").Formula2 = UFCB
#    UFRstart = DS.Range("J5")
#    UFRend = UFRstart.End(c.xlDown)

#    # UFR Cash Balances
#    UFRrange = DS.Range(UFRstart, UFRend)
#    UFRLR = UFRrange.End(c.xlDown).Row
#    # Total Number of Unavialbe Funding Accounts
#    UnavailableFTotalAccounts = str(UFRLR-4)
#    # print(UFRLR)
#    print(UnavailableFTotalAccounts)
#    LOCTotalEnding = str(int(PCRTotalNumberAccounts)+int(UnavailableFTotalAccounts) + 10)
#    print(LOCTotalEnding)

#    UFCBColumn = DS.Columns(10)
#    UFCBColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
   
#    OSdrawdown = DS.Range("L4").Formula2 = OSFilter
#    OSDDCASH = DS.Range("M4").Formula2 = OSCBFilter
#    OSstart = DS.Range("M5")
#    OSEnd = OSstart.End(c.xlDown)
#    OSCBRange = DS.Range(OSstart, OSEnd)
#    OSCBRange.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'


#    PCRCol.AutoFit()
#    PCRrange = DS.Range("G4",PCRCBRow)
#    PCTotal = ExcelApp.WorksheetFunction.Sum(PCRrange)
#    UFTTotal = ExcelApp.WorksheetFunction.Sum(UFRrange)
#    OSTotal = ExcelApp.WorksheetFunction.Sum(OSCBRange)
#    PCRCBtarget2.Value = PCTotal
#    UFTtarget2.Value = UFTTotal
#    OverspentTarget.Value = OSTotal
#    TotalAdjRange = DS.Range(PCRCBtarget2,OverspentTarget)

#    # Total Adjusmtments 
#    TotalAdj = ExcelApp.WorksheetFunction.Sum(TotalAdjRange)
#    CSTAdj = float(TotalAdj * -1) 

#    print(CSTAdj)
#    TAT2.Value = TotalAdj
#    # sheet.range('I5').formula2 = "=FILTER(A:A,(B:B<0)*(C:C=''))"
#    DSCBColumn = DS.Columns(7)
#    DSCBColumn.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

#    new_sheet2.Name = 'LOC CoverSheet'
#    LOC = Drawsheet.Sheets("LOC CoverSheet")
#    LOC.Cells(1, "A").Value = "LOC CoverSheet"
#    LOC.Cells(2, "A").Formula = Today
#    LOC.Cells(9, "A").Value = "Adjustments"
#    LOC.Cells(3, "C").Value = "Sponsor:"
#    LOC.Cells(4, "C").Value = "Period:"
#    LOC.Cells(5, "C").Value = "FM (Desk):"
#    LOCcolumn = LOC.Columns(1)
#    LOCcolumn.AutoFit()
#    LOC.Cells(7, "A").Value = "LOC Report Cash Balance:"
#    LOC.Cells(9, "E").Value = "Comments"
   
#    LOC.Cells(4, "D").Value = '=Sheet1!A4'
#    CommentWidth = LOC.Columns(4)
#    CommentWidth.AutoFit()
#    LOC.Cells(5, "D").Value = '=Sheet1!A6'
   
#    LOC.Cells(3, "D").Value = '=Sheet1!A8'
#    LOC.Cells(7, "D").Value = RCB
#    LOC.Cells(7, "D").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
#    LOC.Cells(9, "D").Value = CSTAdj
#    LOC.Cells(9, "D").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

#    # Positive Cash Balance -> Coversheet Transfer
#    PCRAccounts = CSlength.Copy()
#    CSCBalances = PCashLRange.Copy()
#    CoversheetAccounts = LOC.Range("C11")
#    CoversheetAccounts.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    ExcelApp.CutCopyMode = False

#    UFRAccounts = UFPIDRange.Copy()
#    UFRCoversheet = LOC.Range(CSUFRAccounts)
#    UFRCoversheet.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    ExcelApp.CutCopyMode = False
#    UFRCashbalance = UFRrange.Copy()
#    UFRCBCS = LOC.Range(CSUFRCB)


#    UFRCBCS.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
#    # multiplier = LOC.Range("A3")

   
#    RangeAddress = 'D11:D'+str(LOCTotalEnding)
#    TotalCashNeeded = str(int(LOCTotalEnding) + 2)
#    TotalFundsBeingRequested = str(int(LOCTotalEnding) + 3)
#    Difference = str(int(LOCTotalEnding) + 4)
#    SeniorFMApproval = str(int(LOCTotalEnding) + 6)

#    TCN = LOC.Cells(TotalCashNeeded, "A")
#    TFBR = LOC.Cells(TotalFundsBeingRequested, "A")
#    CSD = LOC.Cells(Difference, "A")

#    TCN.Value = "Total Cash Needed"
#    TFBR.Value = "Total Funds Being Requested"
#    CSD.Value = "Difference"
#    AdjustmentsTotal = '=D7+D9'
#    AdjusmentTotalCoord = LOC.Cells(TotalCashNeeded, "D")
#    TFBRCoordinates = LOC.Cells(TotalFundsBeingRequested, "D")
#    DifferenceCoord = LOC.Cells(Difference, "D")

#    AdjusmentTotalCoord.Formula = AdjustmentsTotal
#    TFBRCoordinates.Value = GTFR
#    DifferenceRange = LOC.Range(AdjusmentTotalCoord,TFBRCoordinates)

#    LOCReconcile = ExcelApp.WorksheetFunction.Sum(DifferenceRange)


#    DifferenceCoord.Value = LOCReconcile
#    DifferenceCoord.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
#    TFBRCoordinates.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'


#    CSTotal = LOC.Range(RangeAddress)
#    for cell in CSTotal:
#       cell.Value = cell.Value * -1
#    CSTotal.NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'

  

def DSFormat():
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   else:
      ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
      ExcelApp.visible = True
      DSF = ExcelApp.Workbooks.Open(SelectedFile)
      DSFsheet = DSF.Sheets(1)
      DSFsheet.Cells.UnMerge()
      DSFsheet.Rows("1:6").Delete()
      # ExcelApp.Application.SendKeys("{TAB}") 
      Drawbalance = DSFsheet.Range("E:E")
      # Drawbalance.Style.WrapText = False
      # DB2 = DSFsheet.Range(Drawbalance, Drawbalance.End(c.xlDown))
      
      shapes = DSFsheet.Shapes
      while shapes.Count > 0:
         shapes.Item(1).Delete()
      dollar = DSFsheet.Cells.Find("$")
      while dollar is not None:
         dollar.Value = dollar.Value.replace("$", "")
         dollar = DSFsheet.Cells.FindNext(dollar)
      # DSFsheet.Range("J:O").EntireColumn.Delete()


      last_row = DSFsheet.Cells(DSFsheet.Rows.Count, 1).End(-4162).Row  # xlUp = -4162
      last_column = DSFsheet.Cells(1, DSFsheet.Columns.Count).End(-4159).Column
      range_to_search = DSFsheet.Range("A3:A{}".format(last_row)) # xlToLeft = -4159
      range_to_sort = DSFsheet.Range(f"A1:J{last_row}")
      range_to_sort.Sort(Key1=DSFsheet.Range("A2"), Order1=1, Orientation=1)
      DSFsheet.Range("D:D").EntireColumn.Insert()
      DSFsheet.Rows("1:1").Insert(-4161)
      DSFsheet.Range("A1").Value = "ASAP Acount"
      DSFsheet.Range("C1").Value = "Account Status"
      DSFsheet.Range("I1").Value = "Available Balance"
      DSFsheet.Range("D1").Value = "Project ID"
      DSFsheet.Range("J1").Value = "Amount Requested"
      
      DSFsheet.Columns("A:A").NumberFormat = "@"
      
      # DSFsheet.Columns(7).Select()
      
      # DSFsheet.Columns(7).TextToColumns(Destination=None)
      # input()
      DSFsheet.Range("I:I").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
      # input()
      # DSFsheet.Columns(4).Delete()
      # DSFsheet.Columns(4).Delete()
      # DSFsheet.Columns(10).Delete()
      # DSFsheet.Columns(10).Delete()
      # DSFsheet.Columns(10).Delete()
      # DSFsheet.Columns(10).Delete()

      for row in range(last_row, 3, -1): 
         cell_value = DSFsheet.Cells(row, 1).Value 
         # print(cell_value) 
         if cell_value and "Cash" in cell_value:  
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1): 
         cell_value = DSFsheet.Cells(row, 1).Value 
         # print(cell_value) 
         if cell_value and "Account" in cell_value:  
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1):  
         cell_value = DSFsheet.Cells(row, 1).Value 
         if cell_value and "Pages" in cell_value:  
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1):  
         cell_value = DSFsheet.Cells(row, 1).Value 
         if cell_value and "Step" in cell_value:  
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1): 
         cell_value = DSFsheet.Cells(row, 1).Value 
         if cell_value and "Recipient" in cell_value:  
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1): 
         cell_value = DSFsheet.Cells(row, 1).Value 
         if cell_value and "Federal" in cell_value:
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1):
         cell_value = DSFsheet.Cells(row, 1).Value 
         if cell_value and "Initiate" in cell_value:
            DSFsheet.Rows(row).Delete()

      for row in range(last_row, 3, -1):
         cell_value = DSFsheet.Cells(row, 1).Value 
         if cell_value and "criteria" in cell_value:  
            DSFsheet.Rows(row).Delete()

      C_row = DSFsheet.Cells(DSFsheet.Rows.Count, 3).End(-4162).Row
      for row in range(C_row, 3, -1): 
         cell_value = DSFsheet.Cells(row, 3).Value 
         # print(cell_value)
         if cell_value and "Payment" in cell_value:  
            DSFsheet.Rows(row).Delete()
      DSFsheet.Columns.AutoFit()   
      DSFsheet.Rows.AutoFit() 
      DSFsheet.Range("I:I").HorizontalAlignment = c.xlCenter
      DSFsheet.Range("I:I").HorizontalAlignment = c.xlLeft

      startrow = DSFsheet.Range("A2")
      ASAPend = startrow.End(c.xlDown)
      ASAPLength = DSFsheet.Range(startrow, ASAPend)
      LastRow = ASAPLength.End(c.xlDown).Row

      # print(LastRow)

      ASAPfinalrow = 'B'+str(LastRow)

      FirstAccount = DSFsheet.Range("B3")

      TotalAccounts = DSFsheet.Range(FirstAccount,ASAPfinalrow)
      AccountFormula = "=MID(A3,6,5)"
      DSFsheet.Range("B3").Formula = AccountFormula
      DSFsheet.Cells(3, "B").AutoFill(TotalAccounts)
      Prin2Text = DSFsheet.Columns(2).Copy()
      BColumn = DSFsheet.Columns(2)
      BColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
      # for col in range(1, 14):  # Columns A through M
           # DSFsheet.Columns(col).AutoFit()
      Status.config(text="DONE!")
      Status.config(foreground="VioletRed1")
      root.event_generate("<<ProcessCompleted>>", when="tail")
      win32com.client.Dispatch("WScript.Shell").AppActivate("ExcelApp")

def THENUMBER():
   DeskCodes = ["3", "4", "5", "6", "7", "8", "9", "A", "B", "C", "G", "H", "I", "J", "L", "M", "O", "R", "V", "Y", "Z"]

   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   else:
      ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
      ExcelApp.visible = True
      Numberrr = ExcelApp.Workbooks.Open(SelectedFile)
      Cash2AR = ExcelApp.Workbooks.Open(SelectedFile2)
      Segrafa = Numberrr.Sheets(1)
      Segrafa.Rows(1).Delete()
      Segrafa.Range("C1").AutoFilter(Field=3, Criteria1="A") 
      Segrafa.Range("F1").AutoFilter(Field=6, Criteria1=DeskCodes, Operator=7)
      Filterdd = Numberrr.Sheets.Add()

      Cash2AR = ExcelApp.Workbooks.Open(SelectedFile2)
      Munyun = Cash2AR.Sheets(1)
      Munyun.Rows("1:5").Delete()
      Everyting = Munyun.UsedRange
      print(Everyting)

      Everyting.CurrentRegion.Sort(Key1=Munyun.Range("A1"), Order1=1, Orientation=1) 

def AZAPPER():
   namecheck = os.path.join(os.getcwd(), OName.get() + ".csv")
   Coverter = os.path.join(os.getcwd(), OName.get() + ".xlsx")
   # print(namecheck)
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   if os.path.exists (namecheck):
      messagebox.showerror('File Already Exist', 'Please Choose a Different FileName')
   
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1")    
         file = SelectedFile
         # file2 = FileSelection
         doc = fitz.open(SelectedFile)
         text_Doc = []
         for pages in range(len(doc)):
            step_size = (pages / len(doc)) * 100
            page = doc[pages]
            text = page.get_text()
            lines = text.split('\n')
            # lines = re.split(r"(\.\d{2}\s*)", text)
            # lines = re.split(r"(?=000\d{6})", text)

            # lines = split(r"(\$\s*-?\d{1,3}(?:,\d{3})*\.\d{2}\s*)", text)

            text_Doc.extend(lines)
            progress_var.set(step_size)
            root.update_idletasks()
            time.sleep(0.1)
         # print(text_Doc)
         ACCOUNT = r"(?P<account>\b\d{20}\b)"
         STATUS  = r"(?P<status>\bOpen\b|\bClosed\b|\bLiquidated\b)"
         AMOUNT = r"(?P<amount>\(?\s*[-+]?\s*\$?\s*(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d{2})?\s*\)?)"
         PATTERN = re.compile(rf"{ACCOUNT}\s*.*?{STATUS}\s*.*?{AMOUNT}",re.IGNORECASE | re.DOTALL | re.VERBOSE)
         rows = []
         for m in PATTERN.finditer(str(text_Doc)):
            account = m.group("account")
            status = m.group("status").capitalize()
            amount = m.group("amount")
            rows.append((account, status, amount))
         print(rows)
         # input()
         csv_file = OName.get()+".csv"
         
         csv_file_path = os.path.join(os.getcwd(), csv_file)
         # os.path.join(os.getcwd(), csv_file)
         # print(csv_file_path)
         with open(csv_file, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Account Code', 'Status', 'Amount'])
            for line in rows:
               writer.writerow(line)

         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         AZAP = ExcelApp.Workbooks.Open(csv_file_path)
         AZAPsheet = AZAP.Sheets(1)
         AZAPsheet.Range("A:A").NumberFormat = '00000000000000000000'
         AZAPsheet.Range("B:B").EntireColumn.Insert()
         AZAPsheet.Range("B1").Value = "Substring"
         AZAPsheet.Range("C:C").EntireColumn.Insert()
         AZAPsheet.Range("E:E").EntireColumn.Insert()
         AZAPsheet.Range("E1").Value = "Project ID"
         AZAPsheet.Range("F:F").EntireColumn.Insert()
         AZAPsheet.Rows(1).Font.Bold = True
         AZAPsheet.Range("D:D").EntireColumn.Font.Bold = True
         AZAPsheet.Range("G:G").EntireColumn.Font.Bold = True
         AZAPsheet.Range("H1").Value = "Draw"
         AZAPsheet.Range("H:H").EntireColumn.Insert()
         AZAPsheet.Range("I1").Value = "Draw Amount"
         AZAPsheet.Range("G:G").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
         AZAPsheet.Range("M1").Value = "Multi Year Accounts"
         AZAPsheet.Range("N1").Value = "ASAP Balance"
         AZAPsheet.Range("O1").Value = "Draw Amount"

         startrow = AZAPsheet.Range("A2")
         ASAPend = startrow.End(c.xlDown)
         ASAPLength = AZAPsheet.Range(startrow, ASAPend)
         LastRow = ASAPLength.End(c.xlDown).Row
         print(LastRow)
         ASAPfinalrow = 'B'+str(LastRow)
         FirstAccount = AZAPsheet.Range("B2")
         TotalAccounts = AZAPsheet.Range(FirstAccount, ASAPfinalrow)
         AccountFormula = "=MID(A2,6,5)"
         AZAPsheet.Range("B2").Formula = AccountFormula
         AZAPsheet.Cells(2, "B").AutoFill(TotalAccounts)

         Prin2Text = AZAPsheet.Columns(2).Copy()
         BColumn = AZAPsheet.Columns(2)
         BColumn.PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)


         AZAPsheet.Columns.AutoFit()
         AZAPsheet.Rows.AutoFit()
         AZAP.SaveAs(Coverter, FileFormat=51)
         os.remove(csv_file_path)
         ImportDraw = messagebox.askyesno("Import Draw", "Do You Want to Import your Drawsheet?")
         if ImportDraw is True:
            DrawSelection = filedialog.askopenfilename(filetypes = [("xlsx files", ".xlsx")])
            DrawFile = Path(DrawSelection)
            # ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
            # ExcelApp.visible = False
            Import = ExcelApp.Workbooks.Open(DrawFile)
            Frontsheet = Import.Sheets("Sheet1")
            Frontsheet.Copy(After=AZAP.Sheets(1))
            Import.Close(SaveChanges=False)
            BaseDraw = AZAP.Sheets("Sheet1")
            BaseDraw.Range("E:E").EntireColumn.Insert()
            SubStringFormula = '=RIGHT(B10,5)'

            used_range = BaseDraw.UsedRange
            column_values = used_range.Columns(1).Value

            for i, cell_value in enumerate(column_values):
               # global found_row
               if cell_value[0] == 'Report Total:':   
                  found_row = i + 1
                  # return found_row
                  print(found_row)

            FirstPID = BaseDraw.Range("E10")
            TotalProjects = AZAPsheet.Range(f"E10:E{found_row}")
            BaseDraw.Range(f"E10:E{found_row}").Formula = SubStringFormula
            BaseDraw.Range(f"E10:E{found_row}").Copy()
            BaseDraw.Range(f"E10:E{found_row}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)

            XlookupFormula = '=XLOOKUP(B2,Sheet1!E:E,Sheet1!A:A,"",0,-1)'
            AZAPsheet.Range("E2").Formula = XlookupFormula
            AZAPsheet.Cells(2, "E").AutoFill(AZAPsheet.Range(f"E2:E{LastRow}"))
            AZAPsheet.Range(f"E2:E{LastRow}").Copy()
            AZAPsheet.Range(f"E2:E{LastRow}").PasteSpecial(Paste=-4163, Operation=-4142, SkipBlanks=False, Transpose=False)
            AZAPsheet.Range("E:E").TextToColumns(Destination=None)

            DrawLookupFormula = '=XLOOKUP(G2,Sheet1!O:O,Sheet1!Q:Q,"",0)'
            AZAPsheet.Range("I2").Formula = DrawLookupFormula
            AZAPsheet.Cells(2, "I").AutoFill(AZAPsheet.Range(f"I2:I{LastRow}"))
            AZAPsheet.Range("I:I").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'  

            FilterFormula = f'=FILTER(UNIQUE(E2:E{LastRow}), COUNTIF(E2:E{LastRow}, UNIQUE(E2:E{LastRow})) > 1)'
            AZAPsheet.Range("M2").Formula2 = FilterFormula

            MultiBalance = '=XLOOKUP(M2,Sheet1!D:D,Sheet1!O:O,"",0)'
            AZAPsheet.Range("N2").Formula = MultiBalance
            
            

            MultiDrawFormula = '=XLOOKUP(N2,Sheet1!O:O,Sheet1!Q:Q,"",0)'
            AZAPsheet.Range("O2").Formula = MultiDrawFormula

            MultiRange = AZAPsheet.UsedRange
            # MYearAccounts = MultiRange.Columns(13).Value
            MASAPstart = AZAPsheet.Range("M2")
            MASAPend = MASAPstart.End(c.xlDown)
            MultiYEAR = AZAPsheet.Range(MASAPstart, MASAPend)
            TotalMulti = MultiYEAR.End(c.xlDown).Row

            AZAPsheet.Cells(2, "N").AutoFill(AZAPsheet.Range(f"N2:N{TotalMulti}"))
            AZAPsheet.Cells(2, "O").AutoFill(AZAPsheet.Range(f"O2:O{TotalMulti}"))
            AZAPsheet.Range(f"N2:N{TotalMulti}").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
            AZAPsheet.Range(f"O2:O{TotalMulti}").NumberFormat = '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
            print(TotalMulti)

         else:
            pass

         # QuikFormat.Close()   
         progress_var.set(100)
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return

def Payroll():
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   else:
      try:
         Status.config(text="RUNNING")
         Status.config(foreground="RoyalBlue1") 
         ExcelApp = win32com.client.dynamic.Dispatch("Excel.Application")
         ExcelApp.visible = True
         P4Format = ExcelApp.Workbooks.Open(SelectedFile)
         progress_var.set(0)
         for sheet_index in range(2, P4Format.Sheets.Count + 1):
            PSheets = P4Format.Sheets(sheet_index)
            progress_var.set(sheet_index)
            root.update_idletasks()
            time.sleep(0.1)
         # sheet_names = [sheet.Name for sheet in P4Format.Sheets]
         # PSheets = P4Format.Sheets(1)
            PSheets.Rows(10).Delete()
            PSheets.Rows(17).Delete()
            PSheets.Cells(17, 'D').Value = "------Salary-------"
            PSheets.Cells(17, 'E').Value = "--Staff Benefits---"
            PSheets.Rows(16).Delete()
            ARange = PSheets.Range('B17:B100')

            for cell in ARange:
               if cell.Value == 'Total EHRA Non-Teaching Salaries':   
                  EHRALabel = cell.Row
                  PSheets.Rows(EHRALabel).Delete()

            for cell in ARange:
               if cell.Value == 'Total SHRA Employee Salaries':
                  SHRALabel = cell.Row
                  PSheets.Rows(SHRALabel).Delete()

            for cell in ARange:
               if cell.Value == 'Total Temporary Wages':
                  TTempW = cell.Row
                  PSheets.Rows(TTempW).Delete()

            for cell in ARange:
               if cell.Value == 'Total Other Staff Benefits':
                  BenefitsLabel = cell.Row
                  PSheets.Rows(BenefitsLabel).Delete()

            BRange = PSheets.Range('C17:C100')

            for cell in BRange:
               if cell.Value == 'EHRA Non-Teaching Salaries':   
                  EHRALabelRed = cell.Row
                  PSheets.Rows(EHRALabelRed).Delete()

            for cell in BRange:
               if cell.Value == 'SHRA Employee Salaries':   
                  SHRALabelRed = cell.Row
                  PSheets.Rows(SHRALabelRed).Delete()

            for cell in BRange: 
               if cell.Value == 'Temporary Wages':
                  TempRed = cell.Row
                  PSheets.Rows(TempRed).Delete()


            for cell in BRange:
               if cell.Value == 'Other Staff Benefits':
                  BenefitsLabelRed = cell.Row
                  PSheets.Rows(BenefitsLabelRed).Delete()

            for cell in ARange:
               if cell.Value == 'University Benefit Charge':
                  FinalTableRow = cell.Row
             
            # print(FinalTableRow)     
                  TableAddress = 'B16:F'+str(FinalTableRow)
                  TableRange = PSheets.Range(TableAddress)
            # print(TableRange)
                  table = PSheets.ListObjects.Add(1, TableRange, 1, 1, None)
         # ResultsPage = PSheets.Sheets(1) 
         progress_var.set(100)
         connection_string = "OLEDB;Provider=Microsoft.Mashup.OleDb.1;Data Source=" + P4Format.FullName
         MCode = '''
         let
            Source = Excel.CurrentWorkbook(),
            #"Filtered Rows" = Table.SelectRows(Source, each [Name] <> "Alldata"),
            #"Removed Other Columns" = Table.SelectColumns(#"Filtered Rows",{"Content"}),
            #"Expanded Content" = Table.ExpandTableColumn(#"Removed Other Columns", "Content", {"Name", "Line Description", "------Salary-------", "--Staff Benefits---", "Total", "Month"}, {"Name", "Line Description", "------Salary-------", "--Staff Benefits---", "Total", "Month"})
         in
            #"Expanded Content"'''
         sheet = P4Format.Worksheets(1)
         P4Data = P4Format.Worksheets(1).Range("A1")
         Alldata = "Alldata"
         P4Format.Queries.Add(Name=Alldata, Formula=MCode)
         # query = P4Format.Queries(Alldata)
         sheet.Activate()
         qt = sheet.QueryTables.Add(Connection=f"OLEDB;Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;Location=Alldata", Destination=sheet.Range("A1"))
         qt.CommandText = f"SELECT * FROM [Alldata]"
         qt.RefreshStyle = 2
         for query in P4Format.Queries:
            print(query.Name)
         time.sleep(2)
         qt.Refresh()
         # P4Format.Save()
         # P4Format.Close()
         Status.config(text="DONE!")
         Status.config(foreground="VioletRed1")
         root.event_generate("<<ProcessCompleted>>", when="tail")
      except:
         messagebox.showerror('OOPS! Syetem Error', 'Sorry! Something Went Wrong: Process Failed!')
         Status.config(text="FAILED!")
         Status.config(foreground="Red")
         PullRequest = messagebox.askyesno("Submit Request?", "Do You Want to Submit a Pull Request?")
         if PullRequest is True:
            webbrowser.open("https://github.com/HoustonAlexander/DATAFORGE/pulls")
         return
      # win32com.client.Dispatch("WScript.Shell").AppActivate("ExcelApp")
      # 
      # sheet.Activate()  
      # win32com.client.Dispatch("WScript.Shell").AppActivate("ExcelApp")
   # WRSsheet.Cells("B52").Value = IDC      
   # Frontsheet = Drawsheet.Sheets("Sheet1")

# cbutton.grid(row=6, column=1, padx=10, pady=10, sticky="ne")

# lbutton = ttk.Button(Wid_frame, text="Exit")
# lbutton.grid(row=7, column=1)

# sbutton = ttk.Button(Wid_frame, text="Start  ", command=start_process_thread, image=playicon, compound="right")
# sbutton.grid(row=8, column=3, padx=5, pady=5, sticky="e")
task_map = {"The Number": THENUMBER, "WireTap": pdf_scan,"RetroActivity": RETROACTIVITIES, "PDFReader": PDFREADER, "Invictus": Invictus, "AZAPPER": AZAPPER, "QuickFormat": QuickFormat, "AutoDraw - USDA NIFA": Autodraw, "AutoDraw - NIH": Autodraw3, "AutoDraw - NSF": Autodraw2, "P4: Payroller": Payroll, "ASAP Drawsheet Formatter": DSFormat }


def start_process_thread():
   selected_option = Letsgo.get()
   func = task_map.get(selected_option)
   # SelectedFile = Path(FileSelection)
   thread = threading.Thread(target=func, daemon=True)
   
   if FileName1.cget("text") == "No File Found":
      messagebox.showerror('Please Choose a File', 'No File Selected!')
   if selected_option == "-Choose Your Process-":
      messagebox.showerror('Please Choose a Process', 'No Process Selected!')
   if is_excel_file_open(FileSelection):
      messagebox.showerror('Selected File is Already Open' , 'Please Close Workbook to Proceed with Action!')
   else:
      thread.start()
      
   print(file_path)

root.bind("<<ProcessCompleted>>", process_completed_handler)
root.resizable(False, False)
root.mainloop()